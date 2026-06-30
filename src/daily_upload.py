"""Daily MTD load report — replicates the manually-maintained
``Daily_Upload_MMDDYYYY.xlsx`` workbook by reading the Alvys Master 2026
xlsx on OneDrive, filtering to month-to-date, and writing a fresh dated
copy back to OneDrive (with email distribution).

Four tabs:

  * **All Loads** — every MTD load (Cancelled excluded), scoped to
    **X-Trux + XFreight** offices only (matches the Power BI report and
    the scorecard email's asset-trucking scope). Grouped by Customer
    Sales Agent, with per-agent subtotals + a grand-total block at the
    bottom that includes the comprehensive "Mileage / Margin / Goal"
    projection.
  * **Customer Loads** — same X-Trux/XFreight scope; direct customers +
    no-customer rows (deadhead / repositioning legs).
  * **Spot Market** — same X-Trux/XFreight scope; broker freight.
  * **X-Linx Loads** — brokerage book. Same per-agent grouped layout,
    plus a brokerage-specific analysis block at the bottom (top
    customers by revenue, top carriers, unique counts, overall margin
    %). The asset-trucking "goal RPM" block is intentionally absent —
    RPM/cost-out doesn't apply to a brokerage with no fleet.

Date filter is **first of the current calendar month → today** based on
the Scheduled Pickup column (matches PBI's monthly bucket).

Open-load empty-mileage estimate: if a load's Load Status isn't
``Completed`` or ``Invoiced`` and its Empty Dispatch Mileage column is
0/blank, substitute :data:`OPEN_EMPTY_ESTIMATE_MI` miles per the user's
spec so the report reads as a fair MTD snapshot even when in-flight
loads haven't been fully accounted yet.

Tunable constants used by the calc blocks live at the top of this
module — change them here, not in the worksheet:

  * :data:`TRUCK_PAY_PER_MI`   — assumed driver pay rate per mile
  * :data:`BREAK_EVEN_RPM`     — RPM needed to cover pay + overhead
  * :data:`GOAL_RPM`           — RPM target for the goal-analysis block
  * :data:`MARGIN_GOAL_MONTHLY` — monthly margin target
  * :data:`NUM_TRUCKS`         — current fleet count
"""
from __future__ import annotations

import io
import logging
import os
import sys
import tempfile
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from src.onedrive_upload import (
    download_file, download_shared_file, ensure_folder, get_token, upload_file,
)
from src.scorecard_email import (
    ALVYS_MASTER_SHARE_URL, compute_qb_pnl, compute_rpm_goal, send_email,
)

log = logging.getLogger("daily_upload")
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
                    datefmt="%H:%M:%S")

CHI_TZ = ZoneInfo("America/Chicago")
OPEN_EMPTY_ESTIMATE_MI = 65
SETTLED_STATUSES = {"completed", "invoiced"}

# OneDrive folder where the post-send "sent-YYYY-MM-DD.txt" marker is
# written. Mirrors the scorecard's pattern: a healthcheck workflow at
# 6:30am CT checks for today's marker and dispatches a recovery run if
# the 5am send dropped silently.
_SENT_MARKER_FOLDER = "DailyUpload"

# --- Goal-analysis tunables (match the manually-maintained sample) --------
TRUCK_PAY_PER_MI    = 1.85
BREAK_EVEN_RPM      = 2.81
GOAL_RPM            = 2.93
MARGIN_GOAL_MONTHLY = 160_000
NUM_TRUCKS          = 17

# X-Linx brokerage runs on Margin %, not rate-per-mile — its tab is built
# around this goal instead of the RPM goal the asset tabs use.
XLINX_MARGIN_GOAL = 0.175

OUTPUT_COLS = [
    "Count", "Customer Sales Agent", "Load #", "Load Status", "Carrier",
    "Customer", "Pick City", "Pick State", "First Pick Status",
    "Drop City", "Drop State", "Last Drop Status",
    "Empty Dispatch Mileage", "Loaded Dispatch Mileage",
    "Customer Revenue", "Driver Rate", "Margin", "Margin %",
]


def _live_goal_rpm(token: str, upn: str, qb_dir: str,
                    alvys_sheets: dict) -> float:
    """Compute the same goal RPM the scorecard shows on page 1 (live
    cost-out: driver pay/mi + office overhead/mi ÷ target OR). Falls back
    to the GOAL_RPM constant when the QB P&L workbook isn't readable.

    Keeping daily_upload and the scorecard on the same goal so the two
    morning emails don't disagree."""
    try:
        pnl_bytes = download_file(token, upn, f"{qb_dir}/QB_ProfitAndLoss.xlsx")
        pnl_sheets = pd.read_excel(io.BytesIO(pnl_bytes), sheet_name=None)
        qb_pnl = compute_qb_pnl(next(iter(pnl_sheets.values())))
        goal = compute_rpm_goal(alvys_sheets, qb_pnl)
        if goal and goal.get("goal_rpm"):
            live = float(goal["goal_rpm"])
            log.info("Live goal RPM from scorecard cost-out: %.4f (constant fallback: %.2f)",
                     live, GOAL_RPM)
            return live
    except Exception as exc:
        log.warning("Could not compute live goal RPM (%s) — using constant %.2f",
                     exc, GOAL_RPM)
    return GOAL_RPM


DIRECT_CUSTOMERS = {
    "berry plastics", "rainbow play", "ascendant", "graham packaging",
    "lewis drug", "viaflex", "billion auto", "billion automotive",
    "dakota potter", "dakota potters", "innovative office",
    "magnum logistics inc - nd", "magnum logistics", "agco",
    "frontier ag", "frontier coop", "sun opta", "sunopta",
    "fortune logistics", "twin cities logistics", "twin city logistics",
    "moc products", "valley queen", "valley queen cheese",
    # ABT Brokerage — XFreight has a co-brokerage agreement in place
    # with ABT, so ABT freight is treated as direct customer freight
    # across all XFreight reports even though the carrier-side name
    # reads "Brokerage". Co-brokered loads count in the Customer Loads
    # tab; never Spot Market.
    "abt brokerage",
}


def _is_direct_customer(name) -> bool:
    n = str(name).strip().lower()
    if not n or n == "nan":
        return False
    segments = [s.strip() for s in n.split("/")]
    return any(seg.startswith(kw) for seg in segments for kw in DIRECT_CUSTOMERS)


def _is_no_customer(name) -> bool:
    n = str(name).strip().lower()
    return n in ("", "nan", "none")


def _is_xtrux_office(name) -> bool:
    """Asset-trucking scope: matches the scorecard's `_alvys_metrics` filter
    so the All / Customer / Spot tabs report the same load universe as page
    1 of the daily brief."""
    n = str(name).strip().lower()
    return any(k in n for k in ("x-trux", "xtrux", "xfreight"))


def _is_xlinx_office(name) -> bool:
    """Brokerage scope — separated onto its own tab with brokerage-style
    analytics (no RPM/goal block, since X-Linx has no fleet)."""
    n = str(name).strip().lower()
    return any(k in n for k in ("x-linx", "xlinx"))


def _pick_source_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cols_lower = {str(c).strip().lower(): c for c in df.columns}
    for c in candidates:
        if c.strip().lower() in cols_lower:
            return cols_lower[c.strip().lower()]
    return None


def _find_col(df: pd.DataFrame, needles: list[str]) -> str | None:
    for needle in needles:
        for c in df.columns:
            if needle in str(c).lower():
                return c
    return None


def _to_naive_dt(series: pd.Series) -> pd.Series:
    d = pd.to_datetime(series, errors="coerce", utc=True)
    try:
        return d.dt.tz_localize(None)
    except (AttributeError, TypeError):
        return pd.to_datetime(series, errors="coerce")


def _resolve_columns(loads: pd.DataFrame) -> dict[str, str | None]:
    mapping = {
        "Customer Sales Agent": ["Customer Sales Agent", "Sales Agent",
                                  "Salesperson", "Sales Rep", "Account Rep"],
        "Load #":               ["Load #", "Load Number", "Load Num", "Load"],
        "Load Status":          ["Load Status", "Status"],
        "Carrier":              ["Carrier", "Carrier Name"],
        "Customer":             ["Customer", "Customer Name"],
        "Pick City":            ["Pick City", "Pickup City", "First Pick City",
                                  "Origin City"],
        "Pick State":           ["Pick State", "Pickup State", "First Pick State",
                                  "Origin State"],
        "First Pick Status":    ["First Pick Status", "Pickup Status",
                                  "Pick Status"],
        "Drop City":            ["Drop City", "Delivery City", "Last Drop City",
                                  "Destination City"],
        "Drop State":           ["Drop State", "Delivery State",
                                  "Last Drop State", "Destination State"],
        "Last Drop Status":     ["Last Drop Status", "Delivery Status",
                                  "Drop Status"],
        "Empty Dispatch Mileage":  ["Empty Dispatch Mileage", "Empty Mileage",
                                     "Empty Miles", "Dead Head Miles", "DH Miles"],
        "Loaded Dispatch Mileage": ["Loaded Dispatch Mileage", "Loaded Mileage",
                                     "Loaded Miles"],
        "Customer Revenue":     ["Customer Revenue", "Revenue", "Total Revenue"],
        "Driver Rate":          ["Driver Rate", "Driver Pay"],
        "Carrier Rate":         ["Carrier Rate"],
    }
    resolved = {out: _pick_source_col(loads, candidates) for out, candidates in mapping.items()}
    missing = [k for k, v in resolved.items() if v is None]
    if missing:
        log.warning("Source columns not found, will be blank: %s", missing)
    return resolved


def _build_normalized(loads: pd.DataFrame, today_chi: pd.Timestamp,
                       trips_df: pd.DataFrame | None = None) -> pd.DataFrame:
    cols = _resolve_columns(loads)
    date_col = _find_col(loads, ["scheduled pickup", "pickup date"])
    if not date_col:
        raise RuntimeError("No date column found in Loads sheet (expected 'Scheduled Pickup').")
    sub = loads.copy()
    dates = _to_naive_dt(sub[date_col])
    mtd_start = pd.Timestamp(today_chi.year, today_chi.month, 1)
    mtd_end   = pd.Timestamp(today_chi.year, today_chi.month, today_chi.day, 23, 59, 59)
    keep = dates.notna() & (dates >= mtd_start) & (dates <= mtd_end)
    sub = sub.loc[keep].copy()
    log.info("Filtered Loads to MTD %s..%s: %d rows", mtd_start.date(), mtd_end.date(), len(sub))

    if "Load Status" in sub.columns:
        before = len(sub)
        sub = sub[sub["Load Status"].astype(str).str.strip().str.lower() != "cancelled"]
        log.info("Dropped %d Cancelled loads (%d remaining)", before - len(sub), len(sub))

    out = pd.DataFrame()
    for out_col, src_col in cols.items():
        out[out_col] = sub[src_col].values if src_col else [None] * len(sub)

    # Carry Office through for tab splitting (X-Trux/XFreight vs X-Linx).
    # Not in OUTPUT_COLS so it never appears in the rendered worksheets.
    office_col = _pick_source_col(sub, ["Office", "Office Name", "Division"])
    out["__Office"] = sub[office_col].values if office_col else ""
    if not office_col:
        log.warning("No Office column found — X-Linx tab will be empty and "
                    "X-Trux scoping cannot be applied.")

    for c in ("Empty Dispatch Mileage", "Loaded Dispatch Mileage",
              "Customer Revenue", "Driver Rate", "Carrier Rate"):
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)

    # Effective load cost, by entity (matches the Power BI Total Load Cost):
    #   • X-Linx (brokerage) loads: Driver Rate + Carrier Rate. The carrier
    #     payout is the real cost, and some brokered loads ALSO carry a
    #     small driver pay — both sum.
    #   • X-Trux / XFreight (asset) loads: Driver Rate ONLY. The Carrier
    #     Rate column on asset loads holds internal revenue allocation
    #     (own fleet, owner-ops) — counting it overstates cost and flips
    #     asset margins negative. Carrier-name matching is NOT reliable
    #     here (owner-op and blank carrier labels), so the rule keys off
    #     the load's Office.
    is_xlinx_row = out["__Office"].apply(_is_xlinx_office)
    xlinx_cr = out["Carrier Rate"].where(is_xlinx_row, 0.0)
    n_cr = int((xlinx_cr > 0).sum())
    out["Driver Rate"] = out["Driver Rate"] + xlinx_cr
    if n_cr:
        log.info("Folded Carrier Rate into load cost on %d X-Linx rows "
                 "(asset loads cost Driver Rate only)", n_cr)

    # The estimators below model COMPANY-DRIVER pay per mile, so they only
    # apply to X-Trux/XFreight asset loads. X-Linx brokered loads price per
    # load (carrier rate), not per mile — estimating their cost from the
    # asset fleet's $/mi both skews the average and invents wrong costs.
    is_asset = out["__Office"].apply(_is_xtrux_office)

    status_lower = out["Load Status"].astype(str).str.strip().str.lower()
    is_open = ~status_lower.isin(SETTLED_STATUSES)
    needs_est = is_asset & is_open & (out["Empty Dispatch Mileage"] <= 0)
    n_est = int(needs_est.sum())
    out.loc[needs_est, "Empty Dispatch Mileage"] = OPEN_EMPTY_ESTIMATE_MI
    if n_est:
        log.info("Set Empty Dispatch Mileage = %d mi for %d open loads (estimate)",
                 OPEN_EMPTY_ESTIMATE_MI, n_est)

    # Open loads carry a $0 Driver Rate until settlement, which makes the
    # margin column read as 100% on those rows and overstates MTD margin.
    # Estimate the rate as total_miles * avg_$/mi using the settled MTD
    # loads' actual rate per mile. Falls back to TRUCK_PAY_PER_MI when
    # there are zero settled loads yet (early in the month).
    total_mi = out["Empty Dispatch Mileage"] + out["Loaded Dispatch Mileage"]
    # avg $/mi from settled ASSET loads only — brokered carrier pay isn't
    # per-mile economics and would inflate the estimate.
    settled = is_asset & (out["Driver Rate"] > 0) & (total_mi > 0)
    settled_mi  = float(total_mi[settled].sum())
    settled_pay = float(out.loc[settled, "Driver Rate"].sum())
    avg_rate = (settled_pay / settled_mi) if settled_mi > 0 else TRUCK_PAY_PER_MI
    # For open loads: sum ALL trip-leg dispatch miles from the Trips sheet
    # (one row per leg) rather than the Loads-row total, which can be 0 or
    # partial mid-trip.  Fall back to the Loads-row total when no Trips entry
    # is found for that load.
    _trip_mi_for_est: pd.Series = total_mi.copy()
    if trips_df is not None and not trips_df.empty:
        _tlc = _find_col(trips_df, ["load #", "load number", "load num"])
        _tl  = _find_col(trips_df, ["loaded dispatch mileage", "loaded mileage", "loaded miles"])
        _te  = _find_col(trips_df, ["empty dispatch mileage", "empty mileage", "empty miles"])
        _tsc = _find_col(trips_df, ["trip status", "status"])
        _lnc = next((c for c in out.columns if str(c).strip() == "Load #"), None)
        if _tlc and _lnc:
            _t = trips_df.copy()
            if _tsc:
                _t = _t[_t[_tsc].astype(str).str.strip().str.lower() != "cancelled"]
            _t["_lid"]   = _t[_tlc].astype(str).str.strip()
            _t["_lmi"]   = pd.to_numeric(_t[_tl], errors="coerce").fillna(0) if _tl else 0
            _t["_emi"]   = pd.to_numeric(_t[_te], errors="coerce").fillna(0) if _te else 0
            _t["_total"] = _t["_lmi"] + _t["_emi"]
            _trip_sum  = _t.groupby("_lid")["_total"].sum()
            _load_ids  = out[_lnc].astype(str).str.strip()
            _mapped    = _load_ids.map(_trip_sum)
            # Only override open loads that have a Trips entry with >0 miles.
            _override  = is_open & _mapped.notna() & (_mapped > 0)
            _trip_mi_for_est = total_mi.copy()
            _trip_mi_for_est[_override] = _mapped[_override]
            n_trip_ov = int(_override.sum())
            if n_trip_ov:
                log.info("Trip-leg miles: %d open loads overridden "
                         "(avg loads-row %.0f mi → avg trip-sum %.0f mi)",
                         n_trip_ov,
                         float(total_mi[_override].mean()),
                         float(_trip_mi_for_est[_override].mean()))

    needs_rate = is_asset & is_open & (out["Driver Rate"] <= 0) & (_trip_mi_for_est > 0)
    n_rate = int(needs_rate.sum())
    if n_rate:
        out.loc[needs_rate, "Driver Rate"] = (_trip_mi_for_est[needs_rate] * avg_rate).round(2)
        source = "settled MTD avg" if settled_mi > 0 else "fallback constant TRUCK_PAY_PER_MI"
        log.info("Estimated Driver Rate for %d open loads at $%.4f/mi (%s; settled: %s mi / $%s pay)",
                 n_rate, avg_rate, source,
                 f"{settled_mi:,.0f}", f"{settled_pay:,.0f}")

    # In-flight load adjustment (per user rule):
    # When Last Drop Status = "open" AND existing Driver Rate is non-zero
    # but suspiciously low (< $200), the load is mid-trip with only a
    # partial driver pay recorded so far. Add 65 empty miles for the
    # return trip and set Driver Rate to the *gap* between the full
    # estimated pay and what's already recorded — so the column shows
    # the additional pay still expected on this load (not the full pay).
    # Note: Margin = Revenue - Driver Rate, so these rows' Margin will
    # be inflated by the existing-pay portion vs settled loads. Treat
    # the per-row margin on these as a planning estimate, not actuals.
    drop_status = out["Last Drop Status"].astype(str).str.strip().str.lower()
    is_drop_open = drop_status.str.contains("open", na=False)
    needs_inflight = (is_asset & is_drop_open
                      & (out["Driver Rate"] > 0) & (out["Driver Rate"] < 200))
    n_inflight = int(needs_inflight.sum())
    if n_inflight:
        out.loc[needs_inflight, "Empty Dispatch Mileage"] = (
            out.loc[needs_inflight, "Empty Dispatch Mileage"] + 65
        )
        new_total = (out.loc[needs_inflight, "Empty Dispatch Mileage"]
                     + out.loc[needs_inflight, "Loaded Dispatch Mileage"])
        existing_rate = out.loc[needs_inflight, "Driver Rate"]
        out.loc[needs_inflight, "Driver Rate"] = (new_total * avg_rate - existing_rate).round(2)
        log.info("In-flight adjustment on %d loads (Last Drop=open, $0<rate<$200): "
                 "+65 empty mi, Driver Rate set to (total * $%.4f/mi) - existing",
                 n_inflight, avg_rate)

    out["Margin"] = out["Customer Revenue"] - out["Driver Rate"]
    out["Margin %"] = (out["Margin"] / out["Customer Revenue"]).where(out["Customer Revenue"] != 0)

    return out


def _split_tabs(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Split the normalized MTD load list into the four report tabs.

    All / Customer / Spot are all X-Trux+XFreight only — matches Power BI
    and the scorecard's asset-trucking scope. X-Linx brokerage gets its
    own tab so it stays out of the asset-truck RPM math. Any rows whose
    Office isn't recognized (typo, blank, new sub-co) are logged but
    dropped from every tab so they don't quietly skew totals."""
    if "__Office" not in df.columns:
        df = df.assign(__Office="")
    xtrux_mask = df["__Office"].apply(_is_xtrux_office)
    xlinx_mask = df["__Office"].apply(_is_xlinx_office)
    other = df.loc[~(xtrux_mask | xlinx_mask)]
    if len(other):
        offices = sorted(set(str(o).strip() for o in other["__Office"]))
        log.info("Dropped %d loads with unrecognized Office (not X-Trux/XFreight/X-Linx): %s",
                 len(other), offices)

    main  = df.loc[xtrux_mask].copy()
    xlinx = df.loc[xlinx_mask].copy()

    is_no_cust = main["Customer"].apply(_is_no_customer)
    is_direct  = main["Customer"].apply(_is_direct_customer)
    customer_mask = is_no_cust | is_direct

    log.info("Tab scope: All=%d, Customer=%d, Spot=%d, X-Linx=%d",
             len(main), int(customer_mask.sum()),
             int((~customer_mask).sum()), len(xlinx))

    return {
        "All Loads":      main,
        "Customer Loads": main.loc[customer_mask].copy(),
        "Spot Market":    main.loc[~customer_mask].copy(),
        "X-Linx Loads":   xlinx,
    }


# ---------------------------------------------------------------------------
# Xlsx writer — replicates the manual workbook's per-agent grouped layout
# ---------------------------------------------------------------------------

# Number formats lifted verbatim from the user's manually-maintained
# sample workbook. Most cells stay on Excel's 'General' format; only
# Margin (and the per-agent / grand-total RPM and $-per-mile cells)
# get the dollar-with-red-on-negative treatment.
_FMT_MARGIN     = '"$"#,##0.00_);[Red]\\("$"#,##0.00\\)'  # data-row margin + RPM-style cells
_FMT_MARGIN_PCT = "0.00%"
_FMT_NUMBER     = "#,##0.00_);[Red](#,##0.00)"            # Total Miles in per-agent calc rows
_FMT_PCT_TENTHS = "0.0%"                                   # DH % in per-agent calc rows
_FMT_PCT        = "0.00%"
_FMT_ACCOUNTING = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
_FMT_NUMBER_NEG = "#,##0.00_);(#,##0.00)"                  # grand-total sum (M-P)

_NUM_FMT = {
    "Margin":   _FMT_MARGIN,
    "Margin %": _FMT_MARGIN_PCT,
}
_HDR_FONT = Font(bold=True)
_BOLD = Font(bold=True)

# Color palette pulled from the sample workbook's "We are at" projection
# block so the report looks visually identical to the one the user has
# been maintaining by hand.
PURPLE_FILL = PatternFill("solid", fgColor="C198E0")  # labels (col A) + header
BLUE_FILL   = PatternFill("solid", fgColor="00B0F0")  # current-period values
GRAY_FILL   = PatternFill("solid", fgColor="A6A6A6")  # break-even / goal values
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")  # tunable constants + flagged rows
RED_FILL    = PatternFill("solid", fgColor="FF0000")  # negative variance to goal


def _write_header(ws, row: int) -> int:
    # Header rows in the user's manual workbook are just bold black text on
    # a white background — no fill. Matching that here.
    for ci, name in enumerate(OUTPUT_COLS, start=1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.font = _HDR_FONT
    return row + 1


def _write_data_row(ws, row: int, count: int, rec: dict) -> int:
    """Emit one load row. Margin / Margin % stay as formulas so a manual
    edit to Revenue or Driver Rate in Excel still recomputes the row."""
    # Per the sample workbook: data cells stay on Excel's General format;
    # only Margin gets the dollar-with-red-on-negative treatment and
    # Margin % is a plain percent. Margin % formula intentionally does NOT
    # wrap in IFERROR — the sample shows #DIV/0! on zero-revenue rows
    # (matches the user's existing workbook behavior).
    values = [
        count,
        rec["Customer Sales Agent"], rec["Load #"], rec["Load Status"],
        rec["Carrier"], rec["Customer"],
        rec["Pick City"], rec["Pick State"], rec["First Pick Status"],
        rec["Drop City"], rec["Drop State"], rec["Last Drop Status"],
        rec["Empty Dispatch Mileage"], rec["Loaded Dispatch Mileage"],
        rec["Customer Revenue"], rec["Driver Rate"],
        f"=O{row}-P{row}",   # Margin = Revenue - Driver Rate
        f"=Q{row}/O{row}",   # Margin % (matches sample's =Q2/O2)
    ]
    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=ci, value=val)
        col = OUTPUT_COLS[ci - 1]
        if col in _NUM_FMT:
            cell.number_format = _NUM_FMT[col]
    return row + 1


def _write_agent_subtotal(ws, row: int, agent: str,
                            data_first: int, data_last: int,
                            margin_centric: bool = False) -> tuple[int, int]:
    """Per-agent subtotal block. Sum row uses plain SUM over the agent's
    data range — open-load filtering happens at generation time (rows are
    dropped from the DataFrame before write), so no in-workbook toggle is
    needed. Returns (next_row, sum_row) so the grand total can reference
    each agent's sum cells."""
    row += 1  # leading blank

    sum_row = row
    rng = lambda c: f"${c}${data_first}:${c}${data_last}"
    # Per sample: sum cells on M-P are plain General format (no commas,
    # no currency); only Q (Margin) gets the red-on-negative dollar fmt.
    ws.cell(row=sum_row, column=13, value=f"=SUM({rng('M')})")
    ws.cell(row=sum_row, column=14, value=f"=SUM({rng('N')})")
    ws.cell(row=sum_row, column=15, value=f"=SUM({rng('O')})")
    ws.cell(row=sum_row, column=16, value=f"=SUM({rng('P')})")
    ws.cell(row=sum_row, column=17,
             value=f"=O{sum_row}-P{sum_row}").number_format = _FMT_MARGIN
    row += 2  # blank

    first = (agent or "").split()[0] if agent else ""
    label = f"{first} Totals" if first else "Totals"
    ws.cell(row=row, column=13, value=label).font = _BOLD
    if margin_centric:
        # X-Linx brokerage: the headline per-agent number is Margin %.
        ws.cell(row=row, column=14, value="Margin %")
        ws.cell(row=row, column=15,
                 value=f"=IFERROR(Q{sum_row}/O{sum_row},0)").number_format = _FMT_MARGIN_PCT
        row += 1
        formulas = (
            ("Revenue",            f"=O{sum_row}",                                          _FMT_ACCOUNTING),
            ("Carrier Pay",        f"=P{sum_row}",                                          _FMT_ACCOUNTING),
            ("Margin",             f"=Q{sum_row}",                                          _FMT_MARGIN),
            ("vs 17.5% Goal",      f"=IFERROR(Q{sum_row}/O{sum_row},0)-{XLINX_MARGIN_GOAL}", _FMT_MARGIN_PCT),
        )
    else:
        ws.cell(row=row, column=14, value="RPM")
        # IFERROR wrapper retained on RPM-style calcs to avoid #DIV/0! cascade
        # when an agent has zero miles (sample has the data so the error doesn't
        # surface there — but defensive).
        ws.cell(row=row, column=15,
                 value=f"=IFERROR(O{sum_row}/(M{sum_row}+N{sum_row}),0)").number_format = _FMT_MARGIN
        row += 1
        formulas = (
            ("Total Miles",                f"=M{sum_row}+N{sum_row}",                                _FMT_NUMBER),
            ("DH %",                       f"=IFERROR(M{sum_row}/(M{sum_row}+N{sum_row}),0)",        _FMT_PCT_TENTHS),
            ("Average Truck Pay per Mile", f"=IFERROR(P{sum_row}/(M{sum_row}+N{sum_row}),0)",        _FMT_MARGIN),
            ("Average Margin Per Mile",    f"=IFERROR(Q{sum_row}/(M{sum_row}+N{sum_row}),0)",        _FMT_MARGIN),
        )
    for lbl, formula, fmt in formulas:
        ws.cell(row=row, column=14, value=lbl)
        ws.cell(row=row, column=15, value=formula).number_format = fmt
        row += 1

    row += 4  # trailing blanks
    return row, sum_row


def _write_grand_total(ws, row: int, agent_sum_rows: list[tuple[str, int, int, int]],
                        data_first: int, data_last: int, total_loads: int,
                        today_chi: pd.Timestamp, include_goal_block: bool,
                        goal_rpm: float, margin_centric: bool = False) -> int:
    """Grand-total + per-agent % table + (All Loads only) goal projection.

    Sum rows are plain SUM-of-agent-sum-cells; per-agent counts are static
    numbers. agent_sum_rows entry =
    (agent_name, data_first, data_last, agent_sum_row)."""
    row = _write_header(ws, row)

    sum_row = row
    def _sum_of_cells(col_letter: str) -> str:
        if not agent_sum_rows:
            return "0"
        parts = [f"{col_letter}{sr[3]}" for sr in agent_sum_rows]
        return "+".join(parts)
    # Sample's grand-total sum uses negative-parens number format on M-P
    # and an accounting dollar on Q (margin).
    ws.cell(row=sum_row, column=1, value=total_loads)
    ws.cell(row=sum_row, column=13, value=f"={_sum_of_cells('M')}").number_format = _FMT_NUMBER_NEG
    ws.cell(row=sum_row, column=14, value=f"={_sum_of_cells('N')}").number_format = _FMT_NUMBER_NEG
    ws.cell(row=sum_row, column=15, value=f"={_sum_of_cells('O')}").number_format = _FMT_NUMBER_NEG
    ws.cell(row=sum_row, column=16, value=f"={_sum_of_cells('P')}").number_format = _FMT_NUMBER_NEG
    ws.cell(row=sum_row, column=17, value=f"=O{sum_row}-P{sum_row}").number_format = _FMT_ACCOUNTING
    for c in (1, 13, 14, 15, 16, 17):
        ws.cell(row=sum_row, column=c).font = _BOLD
    row += 2

    # Per-agent percentage table (cols I/J/K) + headline/goal block (cols N/O).
    ws.cell(row=row, column=9,  value="% of Loads Booked").font = _BOLD
    ws.cell(row=row, column=10, value="% of Revenue").font = _BOLD
    ws.cell(row=row, column=11, value="% of Margin").font = _BOLD
    if margin_centric:
        # X-Linx brokerage: headline number is Margin %, goal block is
        # built on the 17.5% margin goal — no rate-per-mile math.
        ws.cell(row=row, column=14, value="Margin %").font = _BOLD
        pct_cell = ws.cell(row=row, column=15,
                            value=f"=IFERROR(Q{sum_row}/O{sum_row},0)")
        pct_cell.number_format = _FMT_MARGIN_PCT
        pct_cell.font = _BOLD
    else:
        ws.cell(row=row, column=14, value="RPM")
        rpm_cell = ws.cell(row=row, column=15,
                            value=f"=IFERROR(O{sum_row}/(M{sum_row}+N{sum_row}),0)")
        rpm_cell.number_format = _FMT_ACCOUNTING
    row += 1

    if margin_centric:
        goal_block_lines = [
            ("Goal Margin %",              f"={XLINX_MARGIN_GOAL}",                                  _FMT_MARGIN_PCT),
            ("Difference from Goal",       f"=IFERROR(Q{sum_row}/O{sum_row},0)-{XLINX_MARGIN_GOAL}", _FMT_MARGIN_PCT),
            ("Total Revenue",              f"=O{sum_row}",                                           _FMT_ACCOUNTING),
            ("Total Carrier Pay",          f"=P{sum_row}",                                           _FMT_ACCOUNTING),
            ("Total Margin",               f"=Q{sum_row}",                                           _FMT_MARGIN),
            ("Margin at Goal",             f"=O{sum_row}*{XLINX_MARGIN_GOAL}",                       _FMT_ACCOUNTING),
            ("Margin Missed Opportunity",  f"=Q{sum_row}-O{sum_row}*{XLINX_MARGIN_GOAL}",            _FMT_ACCOUNTING),
        ]
    else:
        # Goal-block rows. Per sample these use the accounting-dollar format
        # consistently (no yellow fill on Goal RPM here — the yellow on
        # tunables only shows in the "We are at" projection block below).
        goal_block_lines = [
            # (label, value_formula, fmt)
            ("Goal RPM",                       f"={goal_rpm}",                                                                  _FMT_ACCOUNTING),
            ("Difference from Goal",           f"=O{sum_row}/(M{sum_row}+N{sum_row})-{goal_rpm}",                                _FMT_ACCOUNTING),
            ("% of Difference from Goal",      f"=(O{sum_row}/(M{sum_row}+N{sum_row})-{goal_rpm})/{goal_rpm}",                   _FMT_PCT_TENTHS),
            ("Total Miles",                    f"=M{sum_row}+N{sum_row}",                                                        _FMT_NUMBER),
            ("DH %",                           f"=IFERROR(M{sum_row}/(M{sum_row}+N{sum_row}),0)",                                _FMT_PCT),
            ("Average Truck Pay per Mile",     f"=IFERROR(P{sum_row}/(M{sum_row}+N{sum_row}),0)",                                _FMT_ACCOUNTING),
            ("Average Margin Per Mile",        f"=IFERROR(Q{sum_row}/(M{sum_row}+N{sum_row}),0)",                                _FMT_ACCOUNTING),
            ("Goal Margin Per Mile",           f"={goal_rpm}-{TRUCK_PAY_PER_MI}",                                                _FMT_ACCOUNTING),
            ("Difference from Goal",           f"=O{sum_row}/(M{sum_row}+N{sum_row})-{goal_rpm}",                                _FMT_ACCOUNTING),
            ("Revenue Missed Opportunity",     f"=(O{sum_row}/(M{sum_row}+N{sum_row})-{goal_rpm})*(M{sum_row}+N{sum_row})",      _FMT_ACCOUNTING),
            ("Margin Missed Opportunity",      f"=(O{sum_row}/(M{sum_row}+N{sum_row})-{goal_rpm})*(M{sum_row}+N{sum_row})",      _FMT_ACCOUNTING),
        ]

    n_rows = max(len(agent_sum_rows) + 1, len(goal_block_lines))
    for i in range(n_rows):
        if i < len(agent_sum_rows):
            ag_name, ag_df_first, ag_df_last, ag_sum_row = agent_sum_rows[i]
            first = ag_name.split()[0] if ag_name else ""
            ag_count = ag_df_last - ag_df_first + 1
            ws.cell(row=row, column=8, value=first)
            ws.cell(row=row, column=9,
                     value=f"=IFERROR({ag_count}/$A${sum_row},0)").number_format = "0.00%"
            ws.cell(row=row, column=10,
                     value=f"=IFERROR(O{ag_sum_row}/$O${sum_row},0)").number_format = "0.00%"
            ws.cell(row=row, column=11,
                     value=f"=IFERROR(Q{ag_sum_row}/$Q${sum_row},0)").number_format = "0.00%"
        elif i == len(agent_sum_rows):
            ws.cell(row=row, column=8, value="Total").font = _BOLD
            ws.cell(row=row, column=9,  value=1).number_format = "0.00%"
            ws.cell(row=row, column=10, value=1).number_format = "0.00%"
            ws.cell(row=row, column=11, value=1).number_format = "0.00%"
        if i < len(goal_block_lines):
            label, formula, fmt = goal_block_lines[i]
            ws.cell(row=row, column=14, value=label)
            ws.cell(row=row, column=15, value=formula).number_format = fmt
        row += 1

    if not include_goal_block:
        row += 1
        ws.cell(row=row, column=15, value=f"=$A${sum_row}").font = _BOLD
        row += 1
        ws.cell(row=row, column=14, value="Percentage of Total Loads")
        ws.cell(row=row, column=15, value=1).number_format = "0.00%"
        return row + 1

    # All Loads only — full goal-analysis projection
    row += 1
    ws.cell(row=row, column=15, value=f"=$A${sum_row}").font = _BOLD
    row += 1
    ws.cell(row=row, column=14, value="Percentage of Total Loads")
    ws.cell(row=row, column=15, value=1).number_format = "0.00%"
    row += 2

    ws.cell(row=row, column=2, value="We are at").font = _BOLD
    row += 1

    # Day-of-month and days-in-month are fixed at generation time — these
    # don't need to be formulas. Estimated Mileage uses them with the
    # toggle-aware total miles so the projection scales correctly.
    days_in_month = (pd.Timestamp(today_chi.year, today_chi.month, 1)
                     + pd.offsets.MonthEnd(0)).day
    day_of_month  = today_chi.day or 1
    scale_factor  = days_in_month / day_of_month

    total_mi_ref = f"(M{sum_row}+N{sum_row})"
    rpm_ref      = f"IFERROR(O{sum_row}/{total_mi_ref},0)"
    dh_ref       = f"IFERROR(M{sum_row}/{total_mi_ref},0)"
    est_mi_ref   = f"({total_mi_ref}*{scale_factor})"

    cur_mpm = f"({rpm_ref}-{TRUCK_PAY_PER_MI})"
    be_mpm  = f"({BREAK_EVEN_RPM}-{TRUCK_PAY_PER_MI})"
    gl_mpm  = f"({goal_rpm}-{TRUCK_PAY_PER_MI})"

    cur_marg_est = f"({est_mi_ref}*{cur_mpm})"
    be_marg_est  = f"({est_mi_ref}*{be_mpm})"
    gl_marg_est  = f"({est_mi_ref}*{gl_mpm})"

    cur_est_to_goal = f"({cur_marg_est}-{MARGIN_GOAL_MONTHLY})"
    be_est_to_goal  = f"({be_marg_est}-{MARGIN_GOAL_MONTHLY})"
    gl_est_to_goal  = f"({gl_marg_est}-{MARGIN_GOAL_MONTHLY})"

    cur_short = f"(-({cur_est_to_goal})/IFERROR({cur_mpm},1))"
    be_short  = f"(-({be_est_to_goal})/IFERROR({be_mpm},1))"
    gl_short  = f"(-({gl_est_to_goal})/IFERROR({gl_mpm},1))"

    total_needed_cur = f"({est_mi_ref}+{cur_short})"
    total_needed_be  = f"({est_mi_ref}+{be_short})"
    total_needed_gl  = f"({est_mi_ref}+{gl_short})"

    mi_per_truck    = f"({total_mi_ref}/{NUM_TRUCKS})"
    est_mi_per_truck = f"({est_mi_ref}/{NUM_TRUCKS})"

    need_pt_cur = f"({total_needed_cur}/{NUM_TRUCKS})"
    need_pt_be  = f"({total_needed_be}/{NUM_TRUCKS})"
    need_pt_gl  = f"({total_needed_gl}/{NUM_TRUCKS})"

    short_pt_cur = f"({need_pt_cur}-{est_mi_per_truck})"
    short_pt_be  = f"({need_pt_be}-{est_mi_per_truck})"
    short_pt_gl  = f"({need_pt_gl}-{est_mi_per_truck})"

    trucks_needed_cur = f"IFERROR({total_needed_cur}/{est_mi_per_truck},0)"
    trucks_needed_be  = f"IFERROR({total_needed_be}/{est_mi_per_truck},0)"
    trucks_needed_gl  = f"IFERROR({total_needed_gl}/{est_mi_per_truck},0)"

    proj_rows = [
        ("Dead Head",                   f"={dh_ref}",       f"={dh_ref}",       f"={dh_ref}",       "0.00%",        None),
        ("Trux RPM",                    f"={rpm_ref}",      f"={BREAK_EVEN_RPM}", f"={goal_rpm}",  '"$"#,##0.00',  "tunable_cd"),
        ("Trux Margin Est",             f"={cur_marg_est}", f"={be_marg_est}",  f"={gl_marg_est}",  '"$"#,##0.00',  None),
        ("Truck Miles",                 f"={total_mi_ref}", f"={total_mi_ref}", f"={total_mi_ref}", "#,##0",        None),
        ("Truck Pay",                   f"={TRUCK_PAY_PER_MI}", f"={TRUCK_PAY_PER_MI}", f"={TRUCK_PAY_PER_MI}", '"$"#,##0.00', "tunable_all"),
        ("Truck Margin Per Mile",       f"={cur_mpm}",      f"={be_mpm}",       f"={gl_mpm}",       '"$"#,##0.00',  None),
        ("Estimated Mileage",           f"={est_mi_ref}",   f"={est_mi_ref}",   f"={est_mi_ref}",   "#,##0",        "gray_b"),
        ("Estimated Margin",            f"={cur_marg_est}", f"={be_marg_est}",  f"={gl_marg_est}",  '"$"#,##0.00',  "highlight"),
        ("Margin Needed",               f"={MARGIN_GOAL_MONTHLY}", f"={MARGIN_GOAL_MONTHLY}", f"={MARGIN_GOAL_MONTHLY}", '"$"#,##0.00', "highlight"),
        ("Estimate Margin to Goal",     f"={cur_est_to_goal}", f"={be_est_to_goal}", f"={gl_est_to_goal}", '"$"#,##0.00', "variance"),
        ("Mileage Short Needed for Goal", f"={cur_short}",  f"={be_short}",     f"={gl_short}",     "#,##0",        None),
        ("Total Miles Needed",          f"={total_needed_cur}", f"={total_needed_be}", f"={total_needed_gl}", "#,##0", None),
        ("Number of Trucks",            f"={NUM_TRUCKS}",   f"={NUM_TRUCKS}",   f"={NUM_TRUCKS}",   "0",            "tunable_all"),
        ("Mileage per truck",           f"={mi_per_truck}", f"={mi_per_truck}", f"={mi_per_truck}", "#,##0",        "gray_b"),
        ("Estimated Mileage per truck", f"={est_mi_per_truck}", f"={est_mi_per_truck}", f"={est_mi_per_truck}", "#,##0", "gray_b"),
        ("Mileage Need per truck",      f"={need_pt_cur}",  f"={need_pt_be}",   f"={need_pt_gl}",   "#,##0",        "gray_b"),
        ("Short Mileage per truck",     f"={short_pt_cur}", f"={short_pt_be}",  f"={short_pt_gl}",  "#,##0",        "gray_b"),
        ("Trucks Needed at Estimated Mileage", f"={trucks_needed_cur}", f"={trucks_needed_be}", f"={trucks_needed_gl}", "0.00", "gray_b"),
    ]

    ws.cell(row=row, column=2, value="").fill = PURPLE_FILL
    ws.cell(row=row, column=3, value="Break Even").font = _BOLD
    ws.cell(row=row, column=3).fill = PURPLE_FILL
    ws.cell(row=row, column=4, value="Goal").font = _BOLD
    ws.cell(row=row, column=4).fill = PURPLE_FILL

    for label, cv, bv, gv, fmt, kind in proj_rows:
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = _BOLD
        b_cell = ws.cell(row=row, column=2, value=cv)
        c_cell = ws.cell(row=row, column=3, value=bv)
        d_cell = ws.cell(row=row, column=4, value=gv)
        for cell in (b_cell, c_cell, d_cell):
            cell.number_format = fmt

        label_cell.fill = PURPLE_FILL
        b_cell.fill = BLUE_FILL
        c_cell.fill = GRAY_FILL
        d_cell.fill = GRAY_FILL

        if kind == "tunable_cd":
            c_cell.fill = YELLOW_FILL
            d_cell.fill = YELLOW_FILL
        elif kind == "tunable_all":
            b_cell.fill = YELLOW_FILL
            c_cell.fill = YELLOW_FILL
            d_cell.fill = YELLOW_FILL
        elif kind == "highlight":
            label_cell.fill = YELLOW_FILL
            b_cell.fill = YELLOW_FILL
            c_cell.fill = YELLOW_FILL
            d_cell.fill = YELLOW_FILL
        elif kind == "variance":
            label_cell.fill = YELLOW_FILL
            b_cell.fill = GRAY_FILL
            c_cell.fill = GRAY_FILL
            d_cell.fill = GRAY_FILL
        elif kind == "gray_b":
            b_cell.fill = GRAY_FILL

        row += 1

    return row


def _agents_in_order(df: pd.DataFrame) -> list[str]:
    """Preserve agent order as they appear in the source data; treat empty/nan
    as a single 'Unassigned' bucket. De-dupes while preserving order."""
    seen: list[str] = []
    for v in df["Customer Sales Agent"].astype(str):
        a = v.strip()
        if a.lower() in ("", "nan", "none"):
            a = "Unassigned"
        if a not in seen:
            seen.append(a)
    return seen


def _write_brokerage_analysis(ws, row: int, df: pd.DataFrame) -> int:
    """X-Linx-specific analytics that don't apply to the asset-trucking
    tabs: top customers + top carriers by revenue (with margin %), plus a
    summary block (load/customer/carrier counts, overall margin %).
    Brokerage runs on margin %, not RPM — that's why this block replaces
    the goal-RPM projection on the other tabs."""
    if df.empty:
        return row

    ws.cell(row=row, column=1, value="BROKERAGE ANALYSIS").font = _BOLD
    row += 2

    headers = ["", "Loads", "Revenue", "Carrier Pay", "Margin", "Margin %"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=row, column=ci, value=h).font = _BOLD
    row += 1

    def _top(by: str, n: int = 10) -> pd.DataFrame:
        return (df.groupby(by, dropna=False)
                  .agg(Loads=("Customer Revenue", "size"),
                        Revenue=("Customer Revenue", "sum"),
                        Pay=("Driver Rate", "sum"),
                        Margin=("Margin", "sum"))
                  .sort_values("Revenue", ascending=False)
                  .head(n))

    for label, by in (("Top 10 Customers", "Customer"),
                       ("Top 10 Carriers",  "Carrier")):
        ws.cell(row=row, column=1, value=label).font = _BOLD
        row += 1
        for name, r in _top(by).iterrows():
            rev = float(r["Revenue"] or 0)
            mgn = float(r["Margin"] or 0)
            pct = (mgn / rev) if rev else 0
            display = str(name) if pd.notna(name) and str(name).strip() else "(blank)"
            ws.cell(row=row, column=1, value=display)
            ws.cell(row=row, column=2, value=int(r["Loads"]))
            ws.cell(row=row, column=3, value=rev).number_format = _FMT_ACCOUNTING
            ws.cell(row=row, column=4, value=float(r["Pay"] or 0)).number_format = _FMT_ACCOUNTING
            ws.cell(row=row, column=5, value=mgn).number_format = _FMT_MARGIN
            ws.cell(row=row, column=6, value=pct).number_format = _FMT_MARGIN_PCT
            row += 1
        row += 1

    total_rev = float(df["Customer Revenue"].sum() or 0)
    total_pay = float(df["Driver Rate"].sum() or 0)
    total_mgn = float(df["Margin"].sum() or 0)
    overall_pct = (total_mgn / total_rev) if total_rev else 0
    n_cust = df["Customer"].astype(str).str.strip().replace({"nan": "", "None": ""}).loc[lambda s: s != ""].nunique()
    n_carr = df["Carrier"].astype(str).str.strip().replace({"nan": "", "None": ""}).loc[lambda s: s != ""].nunique()

    ws.cell(row=row, column=1, value="SUMMARY").font = _BOLD
    row += 1
    for label, val, fmt in (
        ("Total Loads",      len(df),     None),
        ("Total Revenue",    total_rev,   _FMT_ACCOUNTING),
        ("Total Carrier Pay", total_pay,  _FMT_ACCOUNTING),
        ("Total Margin",     total_mgn,   _FMT_MARGIN),
        ("Overall Margin %", overall_pct, _FMT_MARGIN_PCT),
        ("Unique Customers", n_cust,      None),
        ("Unique Carriers",  n_carr,      None),
    ):
        ws.cell(row=row, column=1, value=label).font = _BOLD
        c = ws.cell(row=row, column=2, value=val)
        if fmt:
            c.number_format = fmt
        row += 1

    return row


def _autosize_columns(ws, padding: int = 2, min_width: int = 8, max_width: int = 50) -> None:
    """Set each column width to fit its widest non-formula content."""
    col_widths: dict[int, int] = {}
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            if cell.value is None:
                continue
            val = cell.value
            if isinstance(val, str) and val.startswith("="):
                continue
            col_widths[cell.column] = max(col_widths.get(cell.column, 0), len(str(val)))
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(width + padding, max_width)
        )


def _write_tab(ws, df: pd.DataFrame, include_goal_block: bool,
                today_chi: pd.Timestamp, goal_rpm: float,
                brokerage_analysis: bool = False) -> None:
    row = 1
    row = _write_header(ws, row)

    if df.empty:
        ws.cell(row=row, column=1, value="(no MTD loads)")
        return

    agents = _agents_in_order(df)
    agent_sum_rows: list[tuple[str, int, int, int]] = []
    overall_data_first = row
    overall_data_last  = row
    for agent in agents:
        if agent == "Unassigned":
            group = df[df["Customer Sales Agent"].astype(str).str.strip().isin(("", "nan", "None"))]
        else:
            group = df[df["Customer Sales Agent"].astype(str).str.strip() == agent]
        data_first = row
        for count, (_, rec) in enumerate(group.iterrows(), start=1):
            row = _write_data_row(ws, row, count, rec.to_dict())
        data_last = row - 1
        overall_data_last = max(overall_data_last, data_last)
        row, sum_row = _write_agent_subtotal(ws, row, agent, data_first, data_last,
                                              margin_centric=brokerage_analysis)
        agent_sum_rows.append((agent, data_first, data_last, sum_row))

    row = _write_grand_total(ws, row, agent_sum_rows,
                              data_first=overall_data_first,
                              data_last=overall_data_last,
                              total_loads=len(df),
                              today_chi=today_chi,
                              include_goal_block=include_goal_block,
                              goal_rpm=goal_rpm,
                              margin_centric=brokerage_analysis)
    if brokerage_analysis:
        row = _write_brokerage_analysis(ws, row + 2, df)
    _autosize_columns(ws)


def _write_xlsx(tabs: dict[str, pd.DataFrame], file_path: Path,
                 today_chi: pd.Timestamp, goal_rpm: float) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, df in tabs.items():
        ws = wb.create_sheet(title=name)
        _write_tab(ws, df,
                    include_goal_block=(name == "All Loads"),
                    today_chi=today_chi, goal_rpm=goal_rpm,
                    brokerage_analysis=(name == "X-Linx Loads"))
        log.info("Tab %r: %d data rows", name, len(df))
    wb.save(file_path)
    log.info("Wrote %s", file_path)


def _summary_html(tabs: dict[str, pd.DataFrame], file_label: str) -> str:
    parts = ['<div style="font-family:-apple-system,Helvetica,Arial,sans-serif;'
              'font-size:14px;color:#1a1a1a;line-height:1.5;padding:24px;max-width:560px">']
    parts.append('<div style="font-weight:700;letter-spacing:1.5px;font-size:11px;'
                  'color:#c41e2a;text-transform:uppercase;margin-bottom:14px">'
                  'XFreight &middot; Daily MTD Upload</div>')
    parts.append(f"<p style='margin:0 0 12px'>Attached: <b>{file_label}</b> &mdash; "
                  "month-to-date load list grouped by Customer Sales Agent with "
                  "per-agent subtotals and a goal-projection block on All Loads. "
                  "Open / in-flight loads are included with a 65-mi empty-mileage "
                  "estimate so the snapshot reads fairly while loads are still in motion.</p>")
    parts.append("<table cellpadding='6' cellspacing='0' style='border-collapse:collapse;"
                  "border:1px solid #ececec;border-radius:6px;font-size:12.5px;margin:6px 0 16px'>"
                  "<tr style='background:#fafafa;color:#6b6b6b;text-transform:uppercase;"
                  "font-size:10px;letter-spacing:.4px;font-weight:700;'>"
                  "<td>Tab</td><td align='right'>Loads</td>"
                  "<td align='right'>Revenue</td><td align='right'>Margin</td>"
                  "<td align='right'>Margin %</td></tr>")
    for name, df in tabs.items():
        rev = float(df["Customer Revenue"].sum() or 0)
        mgn = float(df["Margin"].sum() or 0)
        pct = (mgn / rev) if rev else 0
        parts.append(
            f"<tr><td style='border-top:1px solid #ececec'>{name}</td>"
            f"<td align='right' style='border-top:1px solid #ececec'>{len(df):,}</td>"
            f"<td align='right' style='border-top:1px solid #ececec'>${rev:,.0f}</td>"
            f"<td align='right' style='border-top:1px solid #ececec'>${mgn:,.0f}</td>"
            f"<td align='right' style='border-top:1px solid #ececec'>{pct*100:.1f}%</td></tr>"
        )
    parts.append("</table>")
    parts.append('<p style="margin:0;color:#6b6b6b;font-size:12px">'
                  f"X-Linx goal: {XLINX_MARGIN_GOAL*100:.1f}% margin. "
                  f"Open loads with no empty mileage on file get a "
                  f"{OPEN_EMPTY_ESTIMATE_MI}-mi estimate. Source: "
                  "<i>Alvys Master2026.xlsx</i> in OneDrive.</p>")
    parts.append("</div>")
    return "".join(parts)


def _pbi_parity_check(loads: pd.DataFrame, normalized: pd.DataFrame,
                       today_chi: pd.Timestamp) -> None:
    """Smoke test: compute Power-BI's standard X-Trux totals plus an
    apples-to-apples 'PBI with open loads' variant, and compare against
    the daily upload's own All Loads totals."""
    log.info("=" * 60)
    log.info("POWER BI PARITY SMOKE TEST  (MTD %s..%s)",
             pd.Timestamp(today_chi.year, today_chi.month, 1).date(),
             today_chi.date())
    log.info("=" * 60)

    date_col = _find_col(loads, ["scheduled pickup", "pickup date"])
    if not date_col:
        log.warning("PBI parity: no date column found, skipping check.")
        return
    sub = loads.copy()
    if "Load Status" in sub.columns:
        sub = sub[sub["Load Status"].astype(str).str.strip().str.lower() != "cancelled"]
    office_col = _pick_source_col(sub, ["Office", "Office Name", "Division"])
    if office_col:
        xtrux_mask = sub[office_col].astype(str).str.strip().str.lower().str.contains(
            "x-trux|xtrux|xfreight", regex=True, na=False)
        sub = sub[xtrux_mask]
        log.info("PBI parity: scoped to X-Trux/XFreight via %r (%d rows)",
                 office_col, len(sub))
    else:
        log.warning("PBI parity: no Office column found — using ALL carriers "
                    "(will overstate vs true PBI X-Trux scope).")
    dates = _to_naive_dt(sub[date_col])
    mtd_start = pd.Timestamp(today_chi.year, today_chi.month, 1)
    mtd_end   = pd.Timestamp(today_chi.year, today_chi.month, today_chi.day, 23, 59, 59)
    sub = sub.loc[dates.notna() & (dates >= mtd_start) & (dates <= mtd_end)].copy()

    loaded_col = _pick_source_col(sub, ["Loaded Miles", "Loaded Mileage"])
    empty_col  = _pick_source_col(sub, ["Empty Miles", "Empty Mileage"])
    rev_col    = _pick_source_col(sub, ["Customer Revenue", "Revenue"])
    rate_col   = _pick_source_col(sub, ["Driver Rate"])

    def _sum(df, col):
        if not col or col not in df.columns:
            return 0
        return float(pd.to_numeric(df[col], errors="coerce").fillna(0).sum())

    def _cost(df):
        return _sum(df, rate_col)

    if rate_col:
        settled = sub[pd.to_numeric(sub[rate_col], errors="coerce").fillna(0) > 0]
    else:
        settled = sub.iloc[0:0]
    pbi_open = sub

    def _block(label: str, df: pd.DataFrame) -> dict:
        loaded = _sum(df, loaded_col)
        empty  = _sum(df, empty_col)
        total  = loaded + empty
        rev    = _sum(df, rev_col)
        pay    = _cost(df)
        rpm    = (rev / total) if total else 0
        return {
            "label": label, "loads": len(df),
            "loaded": loaded, "empty": empty, "total": total,
            "rev": rev, "pay": pay, "rpm": rpm,
        }

    pbi_settled = _block("PBI (settled only)", settled)
    pbi_with_open = _block("PBI + open loads", pbi_open)

    du = {
        "label": "Daily Upload All Loads",
        "loads": len(normalized),
        "loaded": float(normalized["Loaded Dispatch Mileage"].sum()),
        "empty":  float(normalized["Empty Dispatch Mileage"].sum()),
        "rev":    float(normalized["Customer Revenue"].sum()),
        "pay":    float(normalized["Driver Rate"].sum()),
    }
    du["total"] = du["loaded"] + du["empty"]
    du["rpm"]   = (du["rev"] / du["total"]) if du["total"] else 0

    log.info(f"{'METRIC':<22} | {'PBI (settled)':>16} | {'PBI + open':>16} | "
             f"{'Daily Upload':>16} | {'DU vs PBI+open':>16}")
    log.info("-" * 100)
    for key, fmt in (("loads", "{:>16,.0f}"), ("loaded", "{:>16,.0f}"),
                      ("empty", "{:>16,.0f}"), ("total", "{:>16,.0f}"),
                      ("rev", "${:>15,.2f}"), ("pay", "${:>15,.2f}"),
                      ("rpm", "${:>15,.4f}")):
        a = pbi_settled[key]; b = pbi_with_open[key]; c = du[key]
        diff = c - b
        log.info(f"{key:<22} | {fmt.format(a)} | {fmt.format(b)} | "
                 f"{fmt.format(c)} | {fmt.format(diff)}")
    log.info("-" * 100)
    log.info("Note: 'DU vs PBI+open' should be ~0 for loaded/empty/rev/pay if "
              "the daily upload's All Loads tab matches a PBI view that includes "
              "open loads.")
    log.info("=" * 60)


def main() -> int:
    tenant = os.environ["AZURE_TENANT_ID"]
    client = os.environ["AZURE_CLIENT_ID"]
    secret = os.environ["AZURE_CLIENT_SECRET"]
    upn    = os.environ.get("ONEDRIVE_USER_UPN", "jeff@xfreight.net")
    share  = os.environ.get("DAILY_UPLOAD_ALVYS_SHARE_URL", "").strip() or ALVYS_MASTER_SHARE_URL
    if not share:
        raise SystemExit("DAILY_UPLOAD_ALVYS_SHARE_URL is required.")
    out_folder = os.environ.get("DAILY_UPLOAD_FOLDER", "").strip("/")
    qb_dir = os.environ.get("DAILY_UPLOAD_QB_DIR", "QuickBooks").strip("/")
    to_emails = [e.strip()
                 for e in os.environ.get("DAILY_UPLOAD_TO_EMAILS",
                                          "jeff@xfreight.net,Dan@xfreight.net").split(",")
                 if e.strip()]
    token = get_token(tenant, client, secret)

    today_key = pd.Timestamp.now(tz=CHI_TZ).strftime("%Y-%m-%d")
    if (os.environ.get("GITHUB_EVENT_NAME", "").strip() != "workflow_dispatch"
            and os.environ.get("DAILY_UPLOAD_SKIP_IDEMPOTENCY", "").strip() != "1"):
        marker_path = f"{_SENT_MARKER_FOLDER}/sent-{today_key}.txt"
        try:
            download_file(token, upn, marker_path)
            log.info("Today's daily upload was already sent (marker: %s) — skipping.",
                     marker_path)
            return 0
        except Exception as e:
            log.info("No sent marker for %s — proceeding (%s).",
                     today_key, type(e).__name__)

    log.info("Reading Alvys Master 2026 via share URL…")
    workbook_bytes = download_shared_file(token, share)
    sheets = pd.read_excel(io.BytesIO(workbook_bytes), sheet_name=None)
    loads_key = next((k for k in sheets if k.strip().lower() == "loads"), None)
    if not loads_key:
        raise SystemExit(f"No 'Loads' sheet in workbook (have: {list(sheets)})")
    loads = sheets[loads_key]
    log.info("Loads sheet: %d rows, %d cols", len(loads), loads.shape[1])

    trips_key = next((k for k in sheets if k.strip().lower() == "trips"), None)
    trips_df  = sheets[trips_key] if trips_key else None
    if trips_df is not None:
        log.info("Trips sheet: %d rows, %d cols", len(trips_df), trips_df.shape[1])
    else:
        log.info("No Trips sheet found — open-load miles will use Loads-row totals.")

    today_chi = pd.Timestamp.now(tz=CHI_TZ).normalize()
    normalized = _build_normalized(loads, today_chi, trips_df)
    tabs = _split_tabs(normalized)

    _pbi_parity_check(loads, normalized, today_chi)

    goal_rpm = _live_goal_rpm(token, upn, qb_dir, sheets)

    file_label = f"Daily_Upload_{today_chi.strftime('%m%d%Y')}.xlsx"
    with tempfile.TemporaryDirectory() as tmp:
        local_path = Path(tmp) / file_label
        _write_xlsx(tabs, local_path, today_chi, goal_rpm)

        try:
            if out_folder:
                ensure_folder(token, upn, out_folder)
                log.info("Uploading to OneDrive folder %r as %s …", out_folder, file_label)
            else:
                log.info("Uploading to OneDrive root as %s …", file_label)
            upload_file(token, upn, out_folder, file_label, local_path)
        except Exception as exc:
            log.warning("OneDrive upload failed (%s) — sending email anyway.", exc)

        if to_emails:
            with open(local_path, "rb") as fh:
                content_bytes = fh.read()
            send_email(
                token, upn, to_emails,
                f"XFreight Daily MTD Upload — {today_chi.strftime('%b %d, %Y')}",
                _summary_html(tabs, file_label),
                attachments=[{
                    "name": file_label,
                    "content_bytes": content_bytes,
                    "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }],
            )

            try:
                marker_name = f"sent-{today_key}.txt"
                marker_path_local = Path(tmp) / marker_name
                marker_path_local.write_text(
                    f"{today_key}\n{pd.Timestamp.now(tz=CHI_TZ).isoformat()}\n"
                )
                ensure_folder(token, upn, _SENT_MARKER_FOLDER)
                upload_file(token, upn, _SENT_MARKER_FOLDER, marker_name, marker_path_local)
                log.info("Marker written: %s/%s", _SENT_MARKER_FOLDER, marker_name)
            except Exception as exc:
                log.warning("Failed to write 'sent' marker (%s) — healthcheck may resend.", exc)

    log.info("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
