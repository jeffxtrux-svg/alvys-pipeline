"""
Alvys Master Fixer — desktop GUI tool.

Keeps the Alvys Master workbook's Gross Margin column correct:
  • The rule everywhere (Power BI, scorecard, daily upload) is
    Cost = Driver Rate + Carrier Rate, SUMMED. X-Trux loads carry cost in
    Driver Rate, X-Linx brokered loads in Carrier Rate, and some brokered
    loads have BOTH (small driver pay + carrier rate) — both are real cost.
  • This tool loads your TMS export, recomputes
    Gross Margin = Customer Revenue − (Driver Rate + Carrier Rate)
    on every row, and uploads to OneDrive as Alvys Master2026.xlsx.
  • Driver Rate and Carrier Rate are NEVER modified — they stay exactly as
    Alvys exported them, so nothing is double-counted downstream.
  • Same file structure, same tabs, same column names — Power BI unchanged.

Workflow:
  1. Drop your latest TMS export (or download the current OneDrive file)
  2. Click Fix & Upload to OneDrive

Run:
    python -m src.master_fixer_gui
    or double-click "Alvys Master Fixer.command" on the Desktop.
"""
from __future__ import annotations

import io
import logging
import os
import shutil
import tempfile
import threading
from pathlib import Path

import pandas as pd

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND = True
except ImportError:
    _DND = False

import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
except Exception:
    pass

log = logging.getLogger("master_fixer_gui")

# ── palette ───────────────────────────────────────────────────────────────────
BG      = "#1e1e2e"
PANEL   = "#2a2a3e"
CARD    = "#32324a"
ACCENT  = "#7c6af7"
SUCCESS = "#50fa7b"
WARNING = "#f1fa8c"
ERROR   = "#ff5555"
TEXT    = "#f8f8f2"
MUTED   = "#888899"
BORDER  = "#44475a"


# ── helpers ───────────────────────────────────────────────────────────────────

def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    cl = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c in df.columns:
            return c
        if c.lower() in cl:
            return cl[c.lower()]
    return None


def _read_sheets(src: str | Path | bytes) -> dict[str, pd.DataFrame]:
    if isinstance(src, bytes):
        return pd.read_excel(io.BytesIO(src), sheet_name=None, engine="openpyxl")
    path = Path(src)
    if path.suffix.lower() == ".csv":
        return {path.stem: pd.read_csv(path)}
    return pd.read_excel(path, sheet_name=None, engine="openpyxl")


def _count_fixable(sheets: dict[str, pd.DataFrame]) -> int:
    """Rows whose Gross Margin column disagrees with Revenue − (DR + CR)."""
    n = 0
    for df in sheets.values():
        dr_col  = _find_col(df, ["Driver Rate", "DriverRate"])
        cr_col  = _find_col(df, ["Carrier Rate", "CarrierRate", "Sum of Carrier Rate"])
        rev_col = _find_col(df, ["Customer Revenue", "Revenue"])
        gm_col  = _find_col(df, ["Gross Margin", "GrossMargin", "Margin"])
        if not (dr_col or cr_col) or not rev_col or not gm_col:
            continue
        dr  = pd.to_numeric(df[dr_col], errors="coerce").fillna(0) if dr_col else 0
        cr  = pd.to_numeric(df[cr_col], errors="coerce").fillna(0) if cr_col else 0
        rev = pd.to_numeric(df[rev_col], errors="coerce").fillna(0)
        gm  = pd.to_numeric(df[gm_col], errors="coerce").fillna(0)
        n  += int(((rev - (dr + cr)) - gm).abs().gt(0.005).sum())
    return n


def _get_creds() -> dict | None:
    t = os.getenv("AZURE_TENANT_ID")
    c = os.getenv("AZURE_CLIENT_ID")
    s = os.getenv("AZURE_CLIENT_SECRET")
    u = os.getenv("ONEDRIVE_USER_UPN")
    if not all([t, c, s, u]):
        return None
    return {"tenant": t, "client": c, "secret": s, "upn": u,
            "folder": os.getenv("ONEDRIVE_FOLDER_PATH", "")}


def _to_float(val) -> float:
    """Coerce a cell value to float; return 0.0 on None / blank / non-numeric."""
    if val is None or val == "" or val == " ":
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def process_and_upload(
    raw_bytes: bytes,
    target_name: str,
    upload: bool,
    save_path: str | None,
    log_fn,
) -> None:
    """Recompute Gross Margin = Revenue − (Driver Rate + Carrier Rate) in-place
    using openpyxl. DR/CR are never modified; all other cells and data types
    are preserved exactly (no pandas round-trip)."""
    from openpyxl import load_workbook

    log_fn("Loading workbook (openpyxl)…")
    wb = load_workbook(io.BytesIO(raw_bytes))

    total_fixed = 0
    sheets_processed = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue

        # Build header → 1-based column index map from row 1
        headers: dict[str, int] = {}
        for cell in ws[1]:
            if cell.value is not None:
                headers[str(cell.value).strip()] = cell.column

        def find_col_idx(candidates: list[str]) -> int | None:
            hl = {k.lower(): v for k, v in headers.items()}
            for c in candidates:
                if c in headers:
                    return headers[c]
                if c.lower() in hl:
                    return hl[c.lower()]
            return None

        dr_idx  = find_col_idx(["Driver Rate", "DriverRate"])
        cr_idx  = find_col_idx(["Carrier Rate", "CarrierRate", "Sum of Carrier Rate"])
        rev_idx = find_col_idx(["Customer Revenue", "Revenue"])
        gm_idx  = find_col_idx(["Gross Margin", "GrossMargin", "Margin"])

        if not (dr_idx or cr_idx) or not rev_idx or not gm_idx:
            continue  # sheet lacks the needed columns — leave untouched

        n_fixed = 0
        for row in ws.iter_rows(min_row=2):
            dr_num  = _to_float(row[dr_idx - 1].value) if dr_idx else 0.0
            cr_num  = _to_float(row[cr_idx - 1].value) if cr_idx else 0.0
            rev_num = _to_float(row[rev_idx - 1].value)
            gm_cell = row[gm_idx - 1]

            expected = round(rev_num - (dr_num + cr_num), 2)
            if abs(_to_float(gm_cell.value) - expected) > 0.005:
                gm_cell.value = expected
                n_fixed += 1

        if n_fixed:
            log_fn(f"  {sheet_name}: {n_fixed:,} rows — Gross Margin = Revenue − (DR + CR)")
            total_fixed += n_fixed
            sheets_processed += 1
        else:
            log_fn(f"  {sheet_name}: Gross Margin already correct")

    log_fn(f"Fix complete — {total_fixed:,} Gross Margin cells corrected "
           f"across {sheets_processed} sheet(s). Driver/Carrier Rates untouched.")
    log_fn("Saving workbook (all data types preserved)…")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    try:
        wb.save(tmp)

        if save_path:
            shutil.copy(tmp, save_path)
            log_fn(f"✓ Saved to {save_path}")

        if upload:
            creds = _get_creds()
            if not creds:
                log_fn("✗ Missing Azure credentials in .env — cannot upload.")
                return
            log_fn("Connecting to OneDrive…")
            from src.onedrive_upload import get_token, upload_file
            token = get_token(creds["tenant"], creds["client"], creds["secret"])
            log_fn("  Token OK")
            upload_file(token, creds["upn"], creds["folder"], target_name, Path(tmp))
            log_fn(f"✓ Uploaded as '{target_name}'")
    finally:
        try:
            os.unlink(tmp)
        except Exception:
            pass


# ── GUI ───────────────────────────────────────────────────────────────────────

class FileZone(tk.Frame):
    """Drop zone for the TMS export file."""

    def __init__(self, parent, **kw):
        super().__init__(parent, bg=CARD, highlightbackground=BORDER,
                         highlightthickness=1, **kw)
        self._sheets: dict[str, pd.DataFrame] | None = None
        self._raw_bytes: bytes | None = None

        tk.Label(self, text="① TMS Export",
                 font=("SF Pro", 13, "bold"), bg=CARD, fg=TEXT).pack(
            anchor="w", padx=14, pady=(12, 0))
        tk.Label(self,
                 text="Your latest export from Alvys TMS  (Fuel / Loads / Trips).\n"
                      "Drop the file here, browse for it, or download\n"
                      "the current OneDrive file to re-fix it.",
                 font=("SF Pro", 10), bg=CARD, fg=MUTED,
                 justify="left").pack(anchor="w", padx=14, pady=(4, 8))

        btn_row = tk.Frame(self, bg=CARD)
        btn_row.pack(fill="x", padx=14, pady=(0, 6))

        self._dl_btn = tk.Button(
            btn_row, text="⬇  Download from OneDrive",
            font=("SF Pro", 10), bg=ACCENT, fg=TEXT, relief="flat",
            activebackground=ACCENT, cursor="hand2",
            command=self._download)
        self._dl_btn.pack(side="left", ipadx=8, ipady=4)

        tk.Button(btn_row, text="Browse…",
                  font=("SF Pro", 10), bg=PANEL, fg=TEXT, relief="flat",
                  activebackground=BORDER, cursor="hand2",
                  command=self._browse).pack(side="left", padx=(8, 0),
                                             ipadx=8, ipady=4)

        tk.Label(self, text="or drag here",
                 font=("SF Pro", 9), bg=CARD, fg=MUTED).pack()

        self._status = tk.Label(self, text="No file",
                                font=("SF Pro", 10), bg=CARD, fg=MUTED,
                                wraplength=480, justify="left")
        self._status.pack(anchor="w", padx=14, pady=(4, 12))

        if _DND:
            try:
                self.drop_target_register(DND_FILES)
                self.dnd_bind("<<Drop>>", self._on_drop)
            except Exception:
                pass

    @property
    def sheets(self):
        return self._sheets

    @property
    def raw_bytes(self):
        return self._raw_bytes

    def _set_status(self, msg, colour=MUTED):
        self._status.config(text=msg, fg=colour)

    def _browse(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel / CSV", "*.xlsx *.xls *.csv"),
                       ("All files", "*.*")])
        if p:
            self._load(p)

    def _on_drop(self, event):
        raw = event.data.strip()
        if raw.startswith("{") and raw.endswith("}"):
            raw = raw[1:-1]
        self._load(raw.split("} {")[0] if "} {" in raw else raw)

    def _load(self, path: str):
        self._set_status("⏳ Loading…", MUTED)

        def worker():
            try:
                self._raw_bytes = Path(path).read_bytes()
                sheets = _read_sheets(path)
                self._sheets = sheets
                names = list(sheets.keys())
                total = sum(len(d) for d in sheets.values())
                dr_rows = _count_fixable(sheets)
                desc = (f"✓ {Path(path).name}\n"
                        f"[{', '.join(names)}]  {total:,} rows"
                        + (f"  ·  {dr_rows:,} margin cells need fixing" if dr_rows else
                           "  ·  margins already correct"))
                col = SUCCESS if dr_rows else WARNING
                self.after(0, self._set_status, desc, col)
                self.after(0, self.event_generate, "<<FileLoaded>>")
            except Exception as e:
                self.after(0, self._set_status, f"✗ {e}", ERROR)

        threading.Thread(target=worker, daemon=True).start()

    def _download(self):
        creds = _get_creds()
        if not creds:
            messagebox.showerror("Missing credentials",
                                 "Azure credentials not found in .env")
            return
        self._set_status("⏳ Downloading…", MUTED)
        self._dl_btn.config(state="disabled")

        def worker():
            try:
                from src.onedrive_upload import get_token, download_file
                token  = get_token(creds["tenant"], creds["client"], creds["secret"])
                folder = creds["folder"]
                fname  = "Alvys Master2026.xlsx"
                fpath  = f"{folder}/{fname}" if folder else fname
                raw    = download_file(token, creds["upn"], fpath)
                self._raw_bytes = raw
                sheets = _read_sheets(raw)
                self._sheets = sheets
                names  = list(sheets.keys())
                total  = sum(len(d) for d in sheets.values())
                dr_rows = _count_fixable(sheets)
                desc = (f"✓ {fname} (OneDrive)\n"
                        f"[{', '.join(names)}]  {total:,} rows"
                        + (f"  ·  {dr_rows:,} margin cells need fixing" if dr_rows else
                           "  ·  margins already correct"))
                col = SUCCESS if dr_rows else WARNING
                self.after(0, self._set_status, desc, col)
                self.after(0, self.event_generate, "<<FileLoaded>>")
            except Exception as e:
                self.after(0, self._set_status, f"✗ {e}", ERROR)
            finally:
                self.after(0, self._dl_btn.config, {"state": "normal"})

        threading.Thread(target=worker, daemon=True).start()


class App(TkinterDnD.Tk if _DND else tk.Tk):

    def __init__(self):
        super().__init__()
        self.title("Alvys Master Fixer")
        self.configure(bg=BG)
        self.resizable(True, True)
        self.minsize(600, 560)

        self._build_ui()
        logging.basicConfig(level=logging.INFO,
                            format="%(asctime)s %(levelname)-7s %(message)s",
                            datefmt="%H:%M:%S")

    def _build_ui(self):
        # ── header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=PANEL)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Alvys Master Fixer",
                 font=("SF Pro", 16, "bold"), bg=PANEL, fg=TEXT).pack(
            side="left", padx=20, pady=14)
        tk.Label(hdr,
                 text="Recompute Gross Margin = Rev − (DR + CR)  →  Upload to OneDrive",
                 font=("SF Pro", 10), bg=PANEL, fg=MUTED).pack(
            side="left", padx=(0, 20), pady=14)

        # ── file zone ─────────────────────────────────────────────────────────
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=16)

        self._zone = FileZone(body)
        self._zone.pack(fill="x")
        self._zone.bind("<<FileLoaded>>", self._on_file_loaded)

        # ── status line ───────────────────────────────────────────────────────
        self._status_lbl = tk.Label(body, text="Load a file above to begin.",
                                    font=("SF Pro", 10), bg=BG, fg=MUTED,
                                    anchor="w")
        self._status_lbl.pack(fill="x", pady=(10, 0))

        # ── output filename ───────────────────────────────────────────────────
        fn_row = tk.Frame(body, bg=BG)
        fn_row.pack(fill="x", pady=(10, 0))
        tk.Label(fn_row, text="Output filename:", font=("SF Pro", 10),
                 bg=BG, fg=TEXT).pack(side="left")
        self._name_var = tk.StringVar(value="Alvys Master2026.xlsx")
        tk.Entry(fn_row, textvariable=self._name_var,
                 font=("SF Pro", 10), bg=CARD, fg=TEXT,
                 insertbackground=TEXT, relief="flat",
                 width=36).pack(side="left", padx=(8, 0), ipady=4)

        # ── action buttons ────────────────────────────────────────────────────
        btn_row = tk.Frame(body, bg=BG)
        btn_row.pack(fill="x", pady=(12, 0))

        self._upload_btn = tk.Button(
            btn_row, text="↑  Fix & Upload to OneDrive",
            font=("SF Pro", 11, "bold"), bg=ACCENT, fg=TEXT,
            relief="flat", activebackground=ACCENT, cursor="hand2",
            state="disabled", command=self._do_upload)
        self._upload_btn.pack(side="left", ipadx=14, ipady=7)

        self._save_btn = tk.Button(
            btn_row, text="💾  Fix & Save Locally",
            font=("SF Pro", 11), bg=PANEL, fg=TEXT,
            relief="flat", activebackground=BORDER, cursor="hand2",
            state="disabled", command=self._do_save)
        self._save_btn.pack(side="left", padx=(10, 0), ipadx=14, ipady=7)

        tk.Button(btn_row, text="↺  Reset",
                  font=("SF Pro", 11), bg=PANEL, fg=MUTED,
                  relief="flat", activebackground=BORDER, cursor="hand2",
                  command=self._reset).pack(side="right", ipadx=12, ipady=7)

        # ── log ───────────────────────────────────────────────────────────────
        log_frame = tk.Frame(body, bg=CARD, highlightbackground=BORDER,
                             highlightthickness=1)
        log_frame.pack(fill="both", expand=True, pady=(14, 0))

        self._log = tk.Text(log_frame, bg=CARD, fg=TEXT,
                            font=("Menlo", 9), relief="flat",
                            state="disabled", wrap="word",
                            padx=10, pady=8)
        sb = tk.Scrollbar(log_frame, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._log.pack(fill="both", expand=True)

        self._log_msg("Ready — load a file above to begin.")

    # ── events ────────────────────────────────────────────────────────────────

    def _on_file_loaded(self, _event=None):
        self._upload_btn.config(state="normal")
        self._save_btn.config(state="normal")
        self._status_lbl.config(
            text="✓ File loaded.  Click Fix & Upload to OneDrive or Fix & Save Locally.",
            fg=SUCCESS)

    def _log_msg(self, msg: str):
        self._log.config(state="normal")
        self._log.insert("end", msg + "\n")
        self._log.see("end")
        self._log.config(state="disabled")
        log.info(msg)

    def _set_busy(self, busy: bool):
        state = "disabled" if busy else "normal"
        self._upload_btn.config(state=state)
        self._save_btn.config(state=state)

    # ── actions ───────────────────────────────────────────────────────────────

    def _do_upload(self):
        raw = self._zone.raw_bytes
        if not raw:
            return
        name = self._name_var.get().strip() or "Alvys Master2026.xlsx"
        self._set_busy(True)

        def worker():
            try:
                process_and_upload(raw, name,
                                   upload=True, save_path=None,
                                   log_fn=lambda m: self.after(0, self._log_msg, m))
                self.after(0, self._on_done, name, True)
            except Exception as e:
                self.after(0, self._log_msg, f"✗ {e}")
            finally:
                self.after(0, self._set_busy, False)

        threading.Thread(target=worker, daemon=True).start()

    def _do_save(self):
        raw = self._zone.raw_bytes
        if not raw:
            return
        name = self._name_var.get().strip() or "Alvys Master2026.xlsx"
        path = filedialog.asksaveasfilename(
            initialfile=name, defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        self._set_busy(True)

        def worker():
            try:
                process_and_upload(raw, name,
                                   upload=False, save_path=path,
                                   log_fn=lambda m: self.after(0, self._log_msg, m))
                self.after(0, self._on_done, path, False)
            except Exception as e:
                self.after(0, self._log_msg, f"✗ {e}")
            finally:
                self.after(0, self._set_busy, False)

        threading.Thread(target=worker, daemon=True).start()

    def _on_done(self, name: str, uploaded: bool):
        where = "OneDrive" if uploaded else name
        messagebox.showinfo("Done", f"✓ {Path(name).name}\n→ {where}")

    def _reset(self):
        self._zone._sheets = None
        self._zone._raw_bytes = None
        self._zone._set_status("No file", MUTED)
        self._upload_btn.config(state="disabled")
        self._save_btn.config(state="disabled")
        self._status_lbl.config(text="Load a file above to begin.", fg=MUTED)
        self._log.config(state="normal")
        self._log.delete("1.0", "end")
        self._log.config(state="disabled")
        self._log_msg("Ready — load a file above to begin.")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
