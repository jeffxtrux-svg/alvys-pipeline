"""
Settlement Verification Tool
=============================
Drop driver settlement PDFs into settlements_inbox/ then run:

    python -m src.settlement_checker

Compares each settlement against the Alvys pipeline data (OneDrive).

Pay period: Wednesday 3:01 PM CST → following Wednesday 3:00 PM CST
The settlement's own pay period dates drive the Alvys filter window.
"""

import json
import logging
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Driver rate config
# ---------------------------------------------------------------------------
_DRIVER_RATES_FILE = Path(__file__).parent.parent / "driver_rates.json"

def _load_driver_rates() -> dict[str, float]:
    """Load per-mile rates from driver_rates.json. Keys are uppercase driver names."""
    if not _DRIVER_RATES_FILE.exists():
        return {}
    try:
        data = json.loads(_DRIVER_RATES_FILE.read_text())
        return {k.upper().strip(): v for k, v in data.items()
                if not k.startswith("_") and isinstance(v, (int, float))}
    except Exception:
        return {}

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
REPO_ROOT = Path(__file__).parent.parent
SETTLEMENTS_INBOX = REPO_ROOT / "settlements_inbox"
OUTPUT_DIR = REPO_ROOT / "output" / "settlement_reports"

_ONEDRIVE = Path(os.path.expanduser("~/Library/CloudStorage/OneDrive-xfreight.net"))
ALVYS_MASTER_PIPELINE = _ONEDRIVE / "Alvys Master.xlsx"
ALVYS_MASTER_MANUAL = _ONEDRIVE / "XFreight - Claude Working Files/02 - Power BI/Alvys Master.xlsx"

def _pick_alvys_file() -> Path:
    """Use pipeline output if present (most current), otherwise manual master."""
    if ALVYS_MASTER_PIPELINE.exists():
        return ALVYS_MASTER_PIPELINE
    if ALVYS_MASTER_MANUAL.exists():
        return ALVYS_MASTER_MANUAL
    raise FileNotFoundError(
        "Alvys Master.xlsx not found in OneDrive. "
        "Make sure OneDrive is synced."
    )


# ---------------------------------------------------------------------------
# Pay-period helpers
# ---------------------------------------------------------------------------
CST = timezone(timedelta(hours=-6))  # CST = UTC-6 (no DST adjustment — matches Alvys pipeline)


def _parse_alvys_date(val) -> datetime | None:
    """Parse Alvys text date 'MM-DD-YYYY' → datetime at midnight CST."""
    if not val or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return datetime.strptime(str(val).strip(), "%m-%d-%Y").replace(tzinfo=CST)
    except ValueError:
        return None


def week_start_for_date(d: datetime) -> datetime:
    """
    Return the Wednesday 3:01 PM CST that opens the pay week containing d.

    Period boundary: Wed 15:01 CST opens a new week; Wed 15:00 CST closes it.
    A trip delivering on Wednesday before 15:01 belongs to the prior week.
    """
    d_cst = d.astimezone(CST)
    days_back = (d_cst.weekday() - 2) % 7  # steps back to nearest Wednesday
    candidate = (d_cst - timedelta(days=days_back)).replace(
        hour=15, minute=1, second=0, microsecond=0
    )
    # If d falls before that Wednesday's 15:01, step back one week
    if d_cst < candidate:
        candidate -= timedelta(days=7)
    return candidate


def week_end(week_start: datetime) -> datetime:
    """Wednesday 3:00 PM CST, 7 days after week_start (last minute of the period)."""
    return week_start + timedelta(days=7) - timedelta(minutes=1)


def week_label(ws: datetime) -> str:
    we = week_end(ws)
    return f"Wk {ws.strftime('%m/%d')}–{we.strftime('%m/%d/%Y')}"


def parse_iso_dt(val: str | None) -> datetime | None:
    if not val:
        return None
    try:
        # Python 3.11+ handles Z; strip it for older versions
        return datetime.fromisoformat(val.replace("Z", "+00:00"))
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Alvys data loading
# ---------------------------------------------------------------------------

def _count_stops(stops_json: str | None) -> int:
    """Count stops in the Trips.Stops JSON string (excludes first pickup)."""
    if not stops_json or (isinstance(stops_json, float) and pd.isna(stops_json)):
        return 0
    try:
        stops = json.loads(str(stops_json))
        if not isinstance(stops, list):
            return 0
        # Count all non-first-pickup stops (intermediates + deliveries)
        non_pickup = [s for s in stops if s.get("StopType") != "Pickup"]
        return len(non_pickup)
    except (json.JSONDecodeError, TypeError):
        return 0


def _detect_detention_hours(stops_json: str | None) -> float:
    """
    Estimate detention hours from stop dwell times.
    Detention = any stop where dwell > 2 hours (industry standard).
    Returns total detention hours across all stops on the trip.
    """
    if not stops_json or (isinstance(stops_json, float) and pd.isna(stops_json)):
        return 0.0
    try:
        stops = json.loads(str(stops_json))
    except (json.JSONDecodeError, TypeError):
        return 0.0

    total_dtn = 0.0
    DETENTION_THRESHOLD_HRS = 2.0

    for stop in stops:
        arrived = parse_iso_dt(stop.get("ArrivedAt"))
        departed = parse_iso_dt(stop.get("DepartedAt"))
        if not arrived or not departed:
            continue
        dwell = (departed - arrived).total_seconds() / 3600
        if dwell > DETENTION_THRESHOLD_HRS:
            total_dtn += dwell - DETENTION_THRESHOLD_HRS

    return round(total_dtn, 2)


def load_alvys_trips(alvys_path: Path) -> pd.DataFrame:
    """Load Trips sheet and enrich with pay-week and stop metadata."""
    log.info(f"Loading Alvys trips from {alvys_path.name}")
    df = pd.read_excel(alvys_path, sheet_name="Trips").copy()

    df.loc[:, "_delivery_dt"] = df["Scheduled Delivery"].apply(_parse_alvys_date)
    df.loc[:, "_pickup_dt"]   = df["Scheduled Pickup"].apply(_parse_alvys_date)

    df.loc[:, "_week_start"] = df["_delivery_dt"].apply(
        lambda d: week_start_for_date(d) if d is not None else None
    )
    df.loc[:, "_week_label"] = df["_week_start"].apply(
        lambda ws: week_label(ws) if ws is not None else ""
    )
    df.loc[:, "_driver_norm"] = df["Driver 1"].fillna("").str.upper().str.strip()

    # Stop count and estimated detention from the serialized Stops JSON
    if "Stops" in df.columns:
        df.loc[:, "_stop_count"]           = df["Stops"].apply(_count_stops)
        df.loc[:, "_dtn_hours_estimated"]  = df["Stops"].apply(_detect_detention_hours)
    else:
        df.loc[:, "_stop_count"] = 0
        df.loc[:, "_dtn_hours_estimated"] = 0.0

    numeric_cols = [
        "Driver Rate", "Carrier Detention", "Carrier Advances",
        "Carrier Lumper", "Carrier Other Accessorials",
        "Loaded Miles", "Empty Miles", "Total Miles",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df.loc[:, col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)

    log.info(
        f"  {len(df)} trips, {df['Driver 1'].nunique()} drivers, "
        f"{df['_delivery_dt'].notna().sum()} with delivery dates"
    )
    return df


def find_driver_trips(
    name_query: str,
    df: pd.DataFrame,
    period_start: datetime | None = None,
    period_end: datetime | None = None,
) -> tuple[str, pd.DataFrame]:
    """
    Fuzzy-match driver name, then optionally filter to a date window.
    Returns (matched_name, filtered_df).
    """
    q = name_query.upper().strip()

    # Exact match
    exact = df[df["_driver_norm"] == q]
    if not exact.empty:
        matched_name = exact["Driver 1"].iloc[0]
        pool = exact
    else:
        # All query tokens must appear in the driver name
        tokens = q.split()
        mask = df["_driver_norm"].apply(lambda n: all(t in n for t in tokens))
        pool = df[mask]
        if pool.empty:
            log.warning(f"No Alvys driver match for: '{name_query}'")
            return name_query, pd.DataFrame(columns=df.columns)
        matched_name = pool["Driver 1"].iloc[0]
        log.info(f"Fuzzy match: '{name_query}' → '{matched_name}'")

    # Date window filter
    if period_start or period_end:
        if period_start:
            pool = pool[pool["_delivery_dt"].apply(
                lambda d: d is not None and d >= period_start
            )]
        if period_end:
            pool = pool[pool["_delivery_dt"].apply(
                lambda d: d is not None and d <= period_end
            )]

    return matched_name, pool


# ---------------------------------------------------------------------------
# Settlement PDF parsing via Claude API
# ---------------------------------------------------------------------------
EXTRACT_PROMPT = """\
You are extracting structured pay data from a truck driver settlement document.

Return a JSON object with EXACTLY this schema (no extra keys, no comments):

{
  "driver_name": "FIRST LAST",
  "pay_period_start": "YYYY-MM-DD or null",
  "pay_period_end": "YYYY-MM-DD or null",
  "trips": [
    {
      "trip_number": "string or null",
      "load_number": "string or null",
      "pickup_date": "YYYY-MM-DD or null",
      "delivery_date": "YYYY-MM-DD or null",
      "origin": "City, ST or empty string",
      "destination": "City, ST or empty string",
      "loaded_miles": 0.0,
      "empty_miles": 0.0,
      "total_miles": 0.0,
      "mile_pay": 0.0,
      "stop_pay": 0.0,
      "detention_pay": 0.0,
      "layover_pay": 0.0,
      "breakdown_pay": 0.0,
      "other_pay": 0.0,
      "trip_total": 0.0,
      "notes": ""
    }
  ],
  "total_gross_pay": 0.0,
  "total_deductions": 0.0,
  "net_pay": 0.0,
  "deduction_items": [
    {"description": "string", "amount": 0.0}
  ]
}

Rules:
- mile_pay = all loaded + empty mileage pay combined for the trip
- stop_pay = any per-stop or lumper/unloading pay
- layover_pay = overnight layover or sleeper berth pay
- breakdown_pay = any breakdown compensation
- other_pay = anything per-trip not covered above
- Use 0.0 for any numeric field not present
- trip_total should be sum of all trip pay items if not explicitly stated
- Return ONLY the JSON object — no markdown fences, no explanation
"""


def parse_pdf_settlement(file_path: Path) -> dict | None:
    """Extract structured data from a settlement PDF using Claude."""
    try:
        import anthropic
        import base64
    except ImportError:
        log.error("anthropic not installed. Run: pip install anthropic")
        return None

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        log.error("ANTHROPIC_API_KEY not set — cannot parse PDF settlements")
        return None

    log.info(f"Parsing PDF with Claude: {file_path.name}")
    client = anthropic.Anthropic(api_key=api_key)
    b64 = base64.standard_b64encode(file_path.read_bytes()).decode()

    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4096,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
                    },
                    {"type": "text", "text": EXTRACT_PROMPT},
                ],
            }]
        )
        raw = resp.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:].strip()
        return json.loads(raw)
    except Exception as e:
        log.error(f"Claude parse error ({file_path.name}): {e}")
        return None


def parse_excel_settlement(file_path: Path) -> dict | None:
    """
    Minimal Excel/CSV settlement parser.
    Tries to read common layouts; flags for manual review if ambiguous.
    """
    try:
        df = pd.read_csv(file_path) if file_path.suffix.lower() == ".csv" \
            else pd.read_excel(file_path)
        log.info(f"Excel settlement columns: {list(df.columns)}")
        return {
            "driver_name": None,
            "pay_period_start": None,
            "pay_period_end": None,
            "trips": [],
            "total_gross_pay": 0.0,
            "total_deductions": 0.0,
            "net_pay": 0.0,
            "deduction_items": [],
            "_raw_df": df,
            "_needs_manual_review": True,
        }
    except Exception as e:
        log.error(f"Cannot read {file_path.name}: {e}")
        return None


# ---------------------------------------------------------------------------
# Comparison logic
# ---------------------------------------------------------------------------

def _safe_float(v, default: float = 0.0) -> float:
    try:
        f = float(v)
        return f if pd.notna(f) else default
    except (TypeError, ValueError):
        return default


def _parse_period_date(s: str | None) -> datetime | None:
    if not s or s == "null":
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(s).strip(), fmt).replace(tzinfo=CST)
        except ValueError:
            continue
    return None


def build_comparison(
    settlement: dict,
    alvys_driver_trips: pd.DataFrame,
    driver_rates: dict[str, float] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns (trip_rows_df, weekly_summary_df, alvys_only_df).

    trip_rows_df    — one row per settlement trip, with Alvys columns alongside
    weekly_summary_df — aggregated by pay week
    alvys_only_df   — Alvys trips in the period NOT found on the settlement
    """
    s_trips = settlement.get("trips") or []
    period_start = _parse_period_date(settlement.get("pay_period_start"))
    period_end   = _parse_period_date(settlement.get("pay_period_end"))

    # Extend period_end to end-of-day for inclusive filter
    if period_end:
        period_end = period_end.replace(hour=23, minute=59, second=59)

    # Map trip numbers seen on settlement (for "not on settlement" detection)
    matched_alvys_indices: set[int] = set()
    rows = []

    for st in s_trips:
        trip_num = str(st.get("trip_number") or "").strip()
        load_num = str(st.get("load_number") or "").strip()

        # Locate Alvys row by Trip #
        alvys_row = pd.Series(dtype=object)
        idx = None
        if trip_num and not alvys_driver_trips.empty:
            hits = alvys_driver_trips[
                alvys_driver_trips["Trip #"].astype(str).str.strip() == trip_num
            ]
            if not hits.empty:
                idx = hits.index[0]
                alvys_row = hits.iloc[0]
        # Fall back: match on Order # vs load number
        if alvys_row.empty and load_num and not alvys_driver_trips.empty:
            hits = alvys_driver_trips[
                alvys_driver_trips["Order #"].astype(str).str.strip() == load_num
            ]
            if not hits.empty:
                idx = hits.index[0]
                alvys_row = hits.iloc[0]

        if idx is not None:
            matched_alvys_indices.add(idx)

        def a(col: str) -> float:
            return _safe_float(alvys_row.get(col, 0.0) if not alvys_row.empty else 0.0)

        s_loaded   = _safe_float(st.get("loaded_miles"))
        s_empty    = _safe_float(st.get("empty_miles"))
        s_total    = _safe_float(st.get("total_miles")) or (s_loaded + s_empty)
        s_mile_pay = _safe_float(st.get("mile_pay"))
        s_stops    = _safe_float(st.get("stop_pay"))
        s_dtn      = _safe_float(st.get("detention_pay"))
        s_layover  = _safe_float(st.get("layover_pay"))
        s_bkdown   = _safe_float(st.get("breakdown_pay"))
        s_other    = _safe_float(st.get("other_pay"))
        s_trip_tot = _safe_float(st.get("trip_total")) or (
            s_mile_pay + s_stops + s_dtn + s_layover + s_bkdown + s_other
        )

        a_loaded   = a("Loaded Miles")
        a_empty    = a("Empty Miles")
        a_total    = a("Total Miles")

        # Use configured per-mile rate if available, else fall back to Alvys Driver Rate
        driver_name_norm = str(settlement.get("driver_name") or "").upper().strip()
        configured_rate  = (driver_rates or {}).get(driver_name_norm)
        if configured_rate is not None:
            a_mile_pay = round(configured_rate * a_total, 2)
        else:
            a_mile_pay = a("Driver Rate")

        a_stops    = a("Carrier Lumper")
        a_dtn      = a("Carrier Detention")
        a_other    = a("Carrier Other Accessorials")
        a_trip_tot = a_mile_pay + a_stops + a_dtn + a_other
        a_stops_ct = int(alvys_row.get("_stop_count", 0)) if not alvys_row.empty else 0
        a_dtn_hrs  = _safe_float(alvys_row.get("_dtn_hours_estimated", 0)) if not alvys_row.empty else 0.0

        mi_delta       = s_total    - a_total
        mile_pay_delta = s_mile_pay - a_mile_pay
        stops_delta    = s_stops    - a_stops
        dtn_delta      = s_dtn      - a_dtn
        other_delta    = s_other    - a_other
        total_delta    = s_trip_tot - a_trip_tot

        flags = []
        if abs(mi_delta) > 10:
            flags.append(f"MILES ({mi_delta:+.0f})")
        if abs(mile_pay_delta) > 5:
            flags.append(f"MILE PAY ({mile_pay_delta:+.2f})")
        if s_dtn > 0 and a_dtn == 0:
            flags.append(f"DETENTION ${s_dtn:.2f} (not in Alvys)")
        if s_layover > 0:
            flags.append(f"LAYOVER ${s_layover:.2f}")
        if s_bkdown > 0:
            flags.append(f"BREAKDOWN ${s_bkdown:.2f}")
        if s_stops > 0 and a_stops == 0:
            flags.append(f"STOP PAY ${s_stops:.2f} (not in Alvys)")
        if alvys_row.empty:
            flags.append("NOT IN ALVYS")

        # Delivery date → week label
        delivery_str = st.get("delivery_date") or ""
        wk = ""
        if delivery_str and delivery_str != "null":
            try:
                dt = datetime.strptime(delivery_str, "%Y-%m-%d").replace(tzinfo=CST)
                wk = week_label(week_start_for_date(dt))
            except ValueError:
                pass
        if not wk and not alvys_row.empty:
            wk = alvys_row.get("_week_label", "")

        rows.append({
            "Week": wk,
            "Trip #": trip_num or load_num,
            "Pickup": st.get("pickup_date") or "",
            "Delivery": delivery_str,
            "Lane": f"{st.get('origin','')} → {st.get('destination','')}",
            "Alvys Loaded Mi":  a_loaded,
            "Alvys Empty Mi":   a_empty,
            "Alvys Total Mi":   a_total,
            "Settl Loaded Mi":  s_loaded,
            "Settl Empty Mi":   s_empty,
            "Settl Total Mi":   s_total,
            "Mi Δ":             mi_delta,
            "Alvys Mile Pay":   a_mile_pay,
            "Settl Mile Pay":   s_mile_pay,
            "Mile Pay Δ":       mile_pay_delta,
            "Alvys Stops (ct)": a_stops_ct,
            "Alvys Stops $":    a_stops,
            "Settl Stops $":    s_stops,
            "Stops Δ":          stops_delta,
            "Alvys Dtn (est hrs)": a_dtn_hrs,
            "Alvys Dtn $":      a_dtn,
            "Settl Dtn $":      s_dtn,
            "Dtn Δ":            dtn_delta,
            "Settl Layover $":  s_layover,
            "Settl Breakdown $": s_bkdown,
            "Alvys Other $":    a_other,
            "Settl Other $":    s_other,
            "Other Δ":          other_delta,
            "Alvys Trip Total": a_trip_tot,
            "Settl Trip Total": s_trip_tot,
            "Total Δ":          total_delta,
            "Flags":            " | ".join(flags),
            "_matched_alvys":   not alvys_row.empty,
        })

    trip_df = pd.DataFrame(rows)

    # --- Weekly summary ---
    if not trip_df.empty and "Week" in trip_df.columns:
        agg_cols = {
            "Trips": ("Trip #", "count"),
            "Alvys Total Mi":   ("Alvys Total Mi",   "sum"),
            "Settl Total Mi":   ("Settl Total Mi",   "sum"),
            "Mi Δ":             ("Mi Δ",             "sum"),
            "Alvys Mile Pay":   ("Alvys Mile Pay",   "sum"),
            "Settl Mile Pay":   ("Settl Mile Pay",   "sum"),
            "Mile Pay Δ":       ("Mile Pay Δ",       "sum"),
            "Settl Stops $":    ("Settl Stops $",    "sum"),
            "Settl Dtn $":      ("Settl Dtn $",      "sum"),
            "Settl Layover $":  ("Settl Layover $",  "sum"),
            "Settl Breakdown $":("Settl Breakdown $","sum"),
            "Alvys Trip Total": ("Alvys Trip Total", "sum"),
            "Settl Trip Total": ("Settl Trip Total", "sum"),
            "Total Δ":          ("Total Δ",          "sum"),
        }
        weekly_df = trip_df.groupby("Week").agg(**agg_cols).reset_index()
    else:
        weekly_df = pd.DataFrame()

    # --- Alvys trips in the period NOT on the settlement ---
    if not alvys_driver_trips.empty:
        in_period = alvys_driver_trips.copy()
        if period_start:
            in_period = in_period[in_period["_delivery_dt"].apply(
                lambda d: d is not None and d >= period_start
            )]
        if period_end:
            in_period = in_period[in_period["_delivery_dt"].apply(
                lambda d: d is not None and d <= period_end
            )]
        alvys_only = in_period[~in_period.index.isin(matched_alvys_indices)].copy()
        keep = [
            "Trip #", "Order #", "Scheduled Pickup", "Scheduled Delivery",
            "First Stop", "Last Stop", "Loaded Miles", "Empty Miles",
            "Total Miles", "Driver Rate", "_week_label", "_stop_count",
        ]
        alvys_only = alvys_only[[c for c in keep if c in alvys_only.columns]]
        alvys_only = alvys_only.rename(columns={"_week_label": "Week", "_stop_count": "Stop Count"})
    else:
        alvys_only = pd.DataFrame()

    return trip_df, weekly_df, alvys_only


# ---------------------------------------------------------------------------
# Excel report writer
# ---------------------------------------------------------------------------
_H_FILL  = PatternFill("solid", fgColor="1F3864")   # dark blue header
_S_FILL  = PatternFill("solid", fgColor="2E75B6")   # mid blue sub-header
_RED     = PatternFill("solid", fgColor="FFCCCC")
_YELLOW  = PatternFill("solid", fgColor="FFEB9C")
_GREEN   = PatternFill("solid", fgColor="C6EFCE")
_ALT     = PatternFill("solid", fgColor="F2F2F2")
_ORANGE  = PatternFill("solid", fgColor="FCE4D6")   # Alvys-only trips
_H_FONT  = Font(color="FFFFFF", bold=True, size=10)
_B_FONT  = Font(bold=True)
_THIN    = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_MONEY   = '"$"#,##0.00'
_NUM     = '#,##0.0'


def _hrow(ws, row: int, headers: list, widths: list | None = None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = _H_FILL
        c.font = _H_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = _THIN
        if widths and col - 1 < len(widths):
            ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]


def _cell(ws, row: int, col: int, val, fmt: str | None = None,
          fill=None, bold: bool = False):
    c = ws.cell(row=row, column=col, value=val)
    c.border = _THIN
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if bold:
        c.font = _B_FONT
    return c


def _is_delta(col_name: str) -> bool:
    return "Δ" in col_name


def _delta_fill(val):
    if not isinstance(val, (int, float)):
        return None
    if abs(val) < 0.01:
        return _GREEN
    if abs(val) <= 10:
        return _YELLOW
    return _RED


def generate_excel_report(
    driver_name: str,
    settlement: dict,
    trip_df: pd.DataFrame,
    weekly_df: pd.DataFrame,
    alvys_only_df: pd.DataFrame,
    out_path: Path,
):
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1 — Weekly Summary
    # -----------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Weekly Summary"

    # Title block
    def title(row, text, bold=False, size=11):
        c = ws1.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

    title(1, f"Driver Settlement Verification — {driver_name}", bold=True, size=14)
    title(2, (
        f"Settlement Period:  "
        f"{settlement.get('pay_period_start') or '?'}  →  "
        f"{settlement.get('pay_period_end') or '?'}"
    ))
    title(3, (
        f"Settlement Gross: ${settlement.get('total_gross_pay', 0):,.2f}   "
        f"Deductions: ${settlement.get('total_deductions', 0):,.2f}   "
        f"Net Pay: ${settlement.get('net_pay', 0):,.2f}"
    ), bold=True)
    title(4, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    sum_headers = [
        "Week", "Trips",
        "Alvys Mi", "Settl Mi", "Mi Δ",
        "Alvys Mile $", "Settl Mile $", "Mile Pay Δ",
        "Settl Stops $", "Settl Dtn $", "Settl Layover $", "Settl Bkdn $",
        "Alvys Total", "Settl Total", "Total Δ",
    ]
    sum_widths = [24, 7, 12, 12, 9, 14, 14, 12, 13, 13, 14, 13, 14, 14, 12]
    _hrow(ws1, 6, sum_headers, sum_widths)

    if not weekly_df.empty:
        money_idx = {i + 1 for i, h in enumerate(sum_headers) if "$" in h or "Total" in h}
        delta_idx = {i + 1 for i, h in enumerate(sum_headers) if _is_delta(h)}

        for ri, (_, row) in enumerate(weekly_df.iterrows()):
            er = 7 + ri
            alt = _ALT if ri % 2 == 0 else None
            vals = [
                row.get("Week", ""), row.get("Trips", 0),
                row.get("Alvys Total Mi", 0), row.get("Settl Total Mi", 0), row.get("Mi Δ", 0),
                row.get("Alvys Mile Pay", 0), row.get("Settl Mile Pay", 0), row.get("Mile Pay Δ", 0),
                row.get("Settl Stops $", 0), row.get("Settl Dtn $", 0),
                row.get("Settl Layover $", 0), row.get("Settl Breakdown $", 0),
                row.get("Alvys Trip Total", 0), row.get("Settl Trip Total", 0), row.get("Total Δ", 0),
            ]
            for ci, val in enumerate(vals, 1):
                fmt = _MONEY if ci in money_idx else (_NUM if ci in {3, 4, 5} else None)
                fill = _delta_fill(val) if ci in delta_idx else alt
                _cell(ws1, er, ci, val, fmt=fmt, fill=fill)

        # Totals
        tr = 7 + len(weekly_df)
        _cell(ws1, tr, 1, "TOTALS", bold=True)
        sum_map = {
            3: "Alvys Total Mi", 4: "Settl Total Mi", 5: "Mi Δ",
            6: "Alvys Mile Pay", 7: "Settl Mile Pay", 8: "Mile Pay Δ",
            9: "Settl Stops $", 10: "Settl Dtn $", 11: "Settl Layover $", 12: "Settl Breakdown $",
            13: "Alvys Trip Total", 14: "Settl Trip Total", 15: "Total Δ",
        }
        for ci, col_name in sum_map.items():
            if col_name in weekly_df.columns:
                val = weekly_df[col_name].sum()
                fmt = _MONEY if ci >= 6 else _NUM
                fill = _delta_fill(val) if _is_delta(col_name) else None
                _cell(ws1, tr, ci, val, fmt=fmt, fill=fill, bold=True)

    # Legend
    leg_row = (7 + len(weekly_df) + 2) if not weekly_df.empty else 8
    ws1.cell(row=leg_row, column=1, value="Legend:").font = _B_FONT
    for offset, (color, text) in enumerate([
        (_GREEN, "No discrepancy"),
        (_YELLOW, "Small diff (≤$10 / ≤10 mi)"),
        (_RED, "Significant diff (>$10 / >10 mi)"),
        (_ORANGE, "Alvys trip not on settlement"),
    ], 1):
        c = ws1.cell(row=leg_row, column=offset * 2, value=text)
        c.fill = color
        c.border = _THIN

    # -----------------------------------------------------------------------
    # Sheet 2 — Trip Detail
    # -----------------------------------------------------------------------
    if not trip_df.empty:
        ws2 = wb.create_sheet("Trip Detail")

        display = [
            "Week", "Trip #", "Pickup", "Delivery", "Lane",
            "Alvys Total Mi", "Settl Total Mi", "Mi Δ",
            "Alvys Mile Pay", "Settl Mile Pay", "Mile Pay Δ",
            "Alvys Stops (ct)", "Alvys Stops $", "Settl Stops $", "Stops Δ",
            "Alvys Dtn (est hrs)", "Alvys Dtn $", "Settl Dtn $", "Dtn Δ",
            "Settl Layover $", "Settl Breakdown $",
            "Alvys Other $", "Settl Other $", "Other Δ",
            "Alvys Trip Total", "Settl Trip Total", "Total Δ",
            "Flags",
        ]
        widths = [24, 12, 12, 12, 32, 14, 14, 9, 14, 14, 12, 13, 13, 13, 10,
                  18, 12, 12, 9, 14, 15, 13, 13, 10, 15, 15, 12, 40]
        _hrow(ws2, 1, display, widths)

        money_set = {h for h in display if "$" in h or "Total" in h}
        delta_set = {h for h in display if _is_delta(h)}

        for ri, (_, row) in enumerate(trip_df.iterrows()):
            er = 2 + ri
            has_flag = bool(str(row.get("Flags", "")).strip())
            not_in_alvys = not row.get("_matched_alvys", True)
            row_fill = _ORANGE if not_in_alvys else (_ALT if ri % 2 == 0 else None)

            for ci, col in enumerate(display, 1):
                val = row.get(col, "")
                if isinstance(val, float) and pd.isna(val):
                    val = ""
                fmt = _MONEY if col in money_set else (_NUM if "Mi" in col and col not in ("Mi Δ",) else None)
                if col in delta_set:
                    fill = _delta_fill(val)
                elif col == "Flags" and has_flag:
                    fill = _RED
                else:
                    fill = row_fill
                c = _cell(ws2, er, ci, val, fmt=fmt, fill=fill)
                if col == "Flags" and has_flag:
                    c.font = Font(bold=True, color="CC0000")

        # Freeze header
        ws2.freeze_panes = ws2["A2"]

    # -----------------------------------------------------------------------
    # Sheet 3 — Alvys Trips Not on Settlement
    # -----------------------------------------------------------------------
    if not alvys_only_df.empty:
        ws3 = wb.create_sheet("In Alvys — Not on Settlement")
        ws3["A1"] = (
            "These trips appear in Alvys for this driver during the settlement period "
            "but are NOT listed on the settlement. Verify they were included in a prior "
            "or future period, or that the driver was paid."
        )
        ws3["A1"].font = Font(bold=True, color="CC0000")
        ws3.merge_cells("A1:H1")
        ws3.row_dimensions[1].height = 28

        cols = list(alvys_only_df.columns)
        widths3 = [14] * len(cols)
        for i, c in enumerate(cols):
            if "Stop" in c or "Week" in c:
                widths3[i] = 24
            elif "Mi" in c or "Pay" in c or "Rate" in c:
                widths3[i] = 14
        _hrow(ws3, 2, cols, widths3)

        money_set3 = {"Driver Rate"}
        for ri, (_, row) in enumerate(alvys_only_df.iterrows()):
            er = 3 + ri
            fill = _ORANGE if ri % 2 == 0 else PatternFill("solid", fgColor="FDEBD0")
            for ci, col in enumerate(cols, 1):
                val = row.get(col, "")
                fmt = _MONEY if col in money_set3 else None
                _cell(ws3, er, ci, val, fmt=fmt, fill=fill)

    # -----------------------------------------------------------------------
    # Sheet 4 — Deductions
    # -----------------------------------------------------------------------
    ws4 = wb.create_sheet("Deductions")
    ws4["A1"] = "Deductions from Settlement"
    ws4["A1"].font = Font(bold=True, size=12)

    deductions = settlement.get("deduction_items") or []
    if deductions:
        _hrow(ws4, 3, ["Description", "Amount"], [42, 16])
        for ri, item in enumerate(deductions):
            er = 4 + ri
            _cell(ws4, er, 1, item.get("description", ""))
            _cell(ws4, er, 2, item.get("amount", 0.0), fmt=_MONEY)
        tr = 4 + len(deductions)
        _cell(ws4, tr, 1, "TOTAL DEDUCTIONS", bold=True)
        _cell(ws4, tr, 2, settlement.get("total_deductions", 0.0), fmt=_MONEY, bold=True)

    wb.save(out_path)
    log.info(f"Report saved → {out_path}")


# ---------------------------------------------------------------------------
# Consolidated employee-facing report
# ---------------------------------------------------------------------------

def generate_consolidated_report(
    driver_name: str,
    rate_per_mile: float | None,
    all_settlements: list[dict],   # list of {settlement, trip_df, weekly_df}
    out_path: Path,
):
    """
    One Excel file covering all settlement periods for a single driver.
    Designed to be handed to the employee — clear, simple, no jargon.
    """
    wb = Workbook()

    # -----------------------------------------------------------------------
    # Sheet 1 — Pay Summary (one row per week, employee-facing)
    # -----------------------------------------------------------------------
    ws = wb.active
    ws.title = "Pay Summary"

    rate_label = f"${rate_per_mile:.2f}/mi" if rate_per_mile else "Alvys rate"

    # Title block
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"Driver Pay Verification — {driver_name.title()}"
    c.font = Font(bold=True, size=16)
    c.alignment = Alignment(horizontal="left")

    ws.merge_cells("A2:K2")
    ws["A2"].value = f"Contracted mileage rate: {rate_label}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    ws.merge_cells("A3:K3")
    ws["A3"].value = (
        "HOW TO READ THIS REPORT: "
        "'Alvys Miles' = verified miles from dispatch system. "
        "'Expected Pay' = your miles × contracted rate. "
        "'Settlement Amount' = gross amount on your pay stub. "
        "A positive Difference means you were paid MORE than the mileage rate alone."
    )
    ws["A3"].font = Font(size=9, italic=True, color="444444")
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[3].height = 30

    headers = [
        "Pay Week", "# Trips",
        "Alvys Miles", "Rate/Mile", "Expected Mile Pay",
        "Stop Pay", "Detention", "Layover", "Breakdown",
        "Settlement Amount", "Difference", "Status",
    ]
    widths = [26, 8, 13, 11, 18, 12, 12, 12, 12, 18, 14, 10]
    _hrow(ws, 5, headers, widths)

    # Collect all weekly rows across every settlement, keyed by week_start
    week_map: dict[str, dict] = {}

    for item in all_settlements:
        s = item["settlement"]
        wdf = item["weekly_df"]
        tdf = item["trip_df"]
        if wdf is None or wdf.empty:
            continue

        for _, wr in wdf.iterrows():
            wk = wr.get("Week", "")
            if not wk:
                continue
            if wk not in week_map:
                week_map[wk] = {
                    "Week": wk,
                    "Trips": 0,
                    "Alvys Miles": 0.0,
                    "Expected Mile Pay": 0.0,
                    "Stop Pay": 0.0,
                    "Detention": 0.0,
                    "Layover": 0.0,
                    "Breakdown": 0.0,
                    "Settlement Amount": 0.0,
                    "_week_start": None,
                }

            row = week_map[wk]
            row["Trips"]              += int(wr.get("Trips", 0))
            row["Alvys Miles"]        += float(wr.get("Alvys Total Mi", 0))
            row["Expected Mile Pay"]  += float(wr.get("Alvys Mile Pay", 0))
            row["Stop Pay"]           += float(wr.get("Settl Stops $", 0))
            row["Detention"]          += float(wr.get("Settl Dtn $", 0))
            row["Layover"]            += float(wr.get("Settl Layover $", 0))
            row["Breakdown"]          += float(wr.get("Settl Breakdown $", 0))

        # Settlement amount: use gross from settlement header
        # Map it to the week(s) it covers via trip_df delivery dates
        gross = float(s.get("total_gross_pay") or 0)
        if not tdf.empty and "Week" in tdf.columns:
            weeks_in_settl = tdf["Week"].dropna().unique().tolist()
            if len(weeks_in_settl) == 1 and weeks_in_settl[0] in week_map:
                week_map[weeks_in_settl[0]]["Settlement Amount"] += gross

    # Sort weeks chronologically
    def _week_sort_key(wk_label: str):
        try:
            date_part = wk_label.replace("Wk ", "").split("–")[0].strip()
            return datetime.strptime(date_part + "/2026", "%m/%d/%Y")
        except Exception:
            return datetime.min

    sorted_weeks = sorted(week_map.values(), key=lambda r: _week_sort_key(r["Week"]))

    total_miles = total_expected = total_settl = total_stops = total_dtn = total_layover = total_bkdn = 0.0
    total_trips = 0

    for ri, row in enumerate(sorted_weeks):
        er = 6 + ri
        miles    = row["Alvys Miles"]
        expected = row["Expected Mile Pay"]
        settl    = row["Settlement Amount"]
        stops    = row["Stop Pay"]
        dtn      = row["Detention"]
        layover  = row["Layover"]
        bkdn     = row["Breakdown"]
        extras   = stops + dtn + layover + bkdn
        diff     = settl - expected if settl > 0 else None

        # Status
        if settl == 0:
            status = "—"
            status_fill = _ALT
        elif diff is None or abs(diff) < 1:
            status = "OK"
            status_fill = _GREEN
        elif diff >= 0:
            status = "OK +"
            status_fill = _GREEN
        elif diff >= -10:
            status = "~OK"
            status_fill = _YELLOW
        else:
            status = "REVIEW"
            status_fill = _RED

        alt = _ALT if ri % 2 == 0 else None

        vals = [
            row["Week"], row["Trips"],
            miles, f"${rate_per_mile:.2f}" if rate_per_mile else "—",
            expected, stops, dtn, layover, bkdn, settl,
            diff if diff is not None else "—", status,
        ]
        fmts = [None, None, _NUM, None, _MONEY, _MONEY, _MONEY, _MONEY, _MONEY, _MONEY, _MONEY, None]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            fill = status_fill if ci == 12 else alt
            c = _cell(ws, er, ci, val, fmt=fmt, fill=fill)
            if ci == 12:
                c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center" if ci in (1, 2, 4, 12) else "right")

        total_trips   += row["Trips"]
        total_miles   += miles
        total_expected += expected
        total_stops   += stops
        total_dtn     += dtn
        total_layover += layover
        total_bkdn    += bkdn
        if settl > 0:
            total_settl += settl

    # Totals row
    tr = 6 + len(sorted_weeks)
    ws.cell(row=tr, column=1, value="TOTALS").font = Font(bold=True, size=11)
    totals = [
        None, total_trips, total_miles, None, total_expected,
        total_stops, total_dtn, total_layover, total_bkdn, total_settl,
        total_settl - total_expected if total_settl > 0 else None, None,
    ]
    fmts = [None, None, _NUM, None, _MONEY, _MONEY, _MONEY, _MONEY, _MONEY, _MONEY, _MONEY, None]
    for ci, (val, fmt) in enumerate(zip(totals, fmts), 1):
        if val is not None:
            c = _cell(ws, tr, ci, val, fmt=fmt, bold=True)
            c.alignment = Alignment(horizontal="right")

    # -----------------------------------------------------------------------
    # Sheet 2 — Trip Detail (all trips across all settlements, sorted by date)
    # -----------------------------------------------------------------------
    all_trip_rows = []
    for item in all_settlements:
        tdf = item["trip_df"]
        if tdf is not None and not tdf.empty:
            all_trip_rows.append(tdf)

    if all_trip_rows:
        combined = pd.concat(all_trip_rows, ignore_index=True)
        # Remove internal column, sort by delivery date
        if "_matched_alvys" in combined.columns:
            combined = combined.drop(columns=["_matched_alvys"])
        try:
            combined = combined.sort_values("Delivery", ascending=True)
        except Exception:
            pass

        ws2 = wb.create_sheet("Trip Detail")
        display = [
            "Week", "Trip #", "Pickup", "Delivery", "Lane",
            "Alvys Total Mi", "Settl Total Mi",
            "Alvys Mile Pay", "Settl Mile Pay",
            "Settl Stops $", "Settl Dtn $", "Settl Layover $", "Settl Breakdown $",
            "Settl Trip Total", "Flags",
        ]
        widths2 = [24, 12, 12, 12, 32, 14, 14, 14, 14, 13, 13, 14, 16, 16, 40]
        _hrow(ws2, 1, display, widths2)

        money_set2 = {h for h in display if "$" in h or "Pay" in h or "Total" in h}

        for ri, (_, row) in enumerate(combined.iterrows()):
            er = 2 + ri
            has_flag = bool(str(row.get("Flags", "")).strip())
            alt = _ALT if ri % 2 == 0 else None
            for ci, col in enumerate(display, 1):
                val = row.get(col, "")
                if isinstance(val, float) and pd.isna(val):
                    val = ""
                fmt = _MONEY if col in money_set2 else None
                fill = _RED if col == "Flags" and has_flag else alt
                c = _cell(ws2, er, ci, val, fmt=fmt, fill=fill)
                if col == "Flags" and has_flag:
                    c.font = Font(color="CC0000", bold=True)
        ws2.freeze_panes = ws2["A2"]

    wb.save(out_path)
    log.info(f"Consolidated report saved → {out_path}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run():
    try:
        from dotenv import load_dotenv
        load_dotenv(REPO_ROOT / ".env")
    except ImportError:
        pass

    SETTLEMENTS_INBOX.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    files = (
        list(SETTLEMENTS_INBOX.glob("*.pdf"))
        + list(SETTLEMENTS_INBOX.glob("*.xlsx"))
        + list(SETTLEMENTS_INBOX.glob("*.xls"))
        + list(SETTLEMENTS_INBOX.glob("*.csv"))
    )

    if not files:
        print(
            f"\nNo settlement files found in:\n  {SETTLEMENTS_INBOX.resolve()}\n"
            "Drop PDF or Excel settlement files there and re-run.\n"
            "  python -m src.settlement_checker\n"
        )
        return

    print(f"Found {len(files)} settlement file(s).")

    driver_rates = _load_driver_rates()
    if driver_rates:
        print(f"Driver rates loaded: {', '.join(f'{k} @ ${v:.2f}/mi' for k, v in driver_rates.items())}")

    alvys_path = _pick_alvys_file()
    alvys_df = load_alvys_trips(alvys_path)

    # Collect data across all settlements for the consolidated report
    consolidated: dict[str, list] = {}  # driver_name_norm → list of {settlement, trip_df, weekly_df}

    for file_path in files:
        print(f"\n{'=' * 60}")
        print(f"  File:   {file_path.name}")

        if file_path.suffix.lower() == ".pdf":
            settlement = parse_pdf_settlement(file_path)
        else:
            settlement = parse_excel_settlement(file_path)

        if not settlement:
            print(f"  ERROR: Could not parse — skipping.")
            continue

        if settlement.get("_needs_manual_review"):
            print("  NOTE: Excel settlements need manual review — partial output only.")

        driver_name = settlement.get("driver_name") or "UNKNOWN"
        print(f"  Driver: {driver_name}")
        print(f"  Period: {settlement.get('pay_period_start')} → {settlement.get('pay_period_end')}")
        print(f"  Trips on settlement: {len(settlement.get('trips') or [])}")
        print(f"  Settlement gross: ${settlement.get('total_gross_pay', 0):,.2f}")

        p_start = _parse_period_date(settlement.get("pay_period_start"))
        p_end   = _parse_period_date(settlement.get("pay_period_end"))
        matched_name, driver_trips = find_driver_trips(driver_name, alvys_df, p_start, p_end)
        print(f"  Alvys match: '{matched_name}' — {len(driver_trips)} trips in period")

        trip_df, weekly_df, alvys_only_df = build_comparison(settlement, driver_trips, driver_rates)

        safe_name  = driver_name.replace(" ", "_").replace("/", "-")
        raw_end    = settlement.get("pay_period_end") or datetime.now().strftime("%Y-%m-%d")
        # Normalize date to YYYYMMDD — strip any separators and non-digits
        parsed_end = _parse_period_date(str(raw_end))
        period_tag = parsed_end.strftime("%Y%m%d") if parsed_end else "".join(c for c in str(raw_end) if c.isdigit())
        out_file   = OUTPUT_DIR / f"Settlement_{safe_name}_{period_tag}.xlsx"

        generate_excel_report(matched_name, settlement, trip_df, weekly_df, alvys_only_df, out_file)

        # Accumulate for consolidated report
        key = matched_name.upper().strip()
        if key not in consolidated:
            consolidated[key] = []
        consolidated[key].append({
            "settlement": settlement,
            "trip_df": trip_df,
            "weekly_df": weekly_df,
            "matched_name": matched_name,
        })

        # Console summary
        if not weekly_df.empty:
            print("\n  Weekly Breakdown:")
            print(f"  {'Week':<28} {'Trips':>5}  {'Alvys Mi':>9}  {'Settl Mi':>9}  "
                  f"{'Alvys $':>10}  {'Settl $':>10}  {'Δ':>10}  Status")
            print("  " + "-" * 95)
            for _, wr in weekly_df.iterrows():
                delta = wr.get("Total Δ", 0)
                status = "OK" if abs(delta) < 5 else ("SMALL DIFF" if abs(delta) <= 25 else "REVIEW")
                print(
                    f"  {wr['Week']:<28} {int(wr.get('Trips',0)):>5}  "
                    f"{wr.get('Alvys Total Mi',0):>9,.0f}  {wr.get('Settl Total Mi',0):>9,.0f}  "
                    f"${wr.get('Alvys Trip Total',0):>9,.2f}  ${wr.get('Settl Trip Total',0):>9,.2f}  "
                    f"${delta:>+9,.2f}  {status}"
                )
            total_delta = weekly_df["Total Δ"].sum()
            print("  " + "-" * 95)
            print(
                f"  {'TOTAL':<28} {int(weekly_df['Trips'].sum()):>5}  "
                f"{weekly_df['Alvys Total Mi'].sum():>9,.0f}  {weekly_df['Settl Total Mi'].sum():>9,.0f}  "
                f"${weekly_df['Alvys Trip Total'].sum():>9,.2f}  ${weekly_df['Settl Trip Total'].sum():>9,.2f}  "
                f"${total_delta:>+9,.2f}"
            )

        if not trip_df.empty:
            flagged = trip_df[trip_df["Flags"].str.strip() != ""]
            if not flagged.empty:
                print(f"\n  ⚠  {len(flagged)} trip(s) flagged:")
                for _, fr in flagged.iterrows():
                    print(f"     Trip {fr['Trip #']}:  {fr['Flags']}")

        if not alvys_only_df.empty:
            print(f"\n  ⚠  {len(alvys_only_df)} Alvys trip(s) NOT found on settlement:")
            for _, ar in alvys_only_df.iterrows():
                print(
                    f"     Trip {ar.get('Trip #','')}  "
                    f"{ar.get('Scheduled Pickup','')} → {ar.get('Scheduled Delivery','')}  "
                    f"{ar.get('Total Miles',0):.0f} mi  ${ar.get('Driver Rate',0):.2f}"
                )

        print(f"\n  Report → {out_file}")

    # -----------------------------------------------------------------------
    # Generate one consolidated report per driver
    # -----------------------------------------------------------------------
    if consolidated:
        print(f"\n{'=' * 60}")
        print("Generating consolidated reports...")
        for driver_key, items in consolidated.items():
            matched_name = items[0]["matched_name"]
            rate = driver_rates.get(driver_key)
            safe  = matched_name.replace(" ", "_").replace("/", "-")
            cpath = OUTPUT_DIR / f"CONSOLIDATED_{safe}.xlsx"
            generate_consolidated_report(matched_name, rate, items, cpath)
            print(f"  {matched_name}: {cpath.name}")

    print(f"\n{'=' * 60}\nDone.\n")


if __name__ == "__main__":
    from dotenv import load_dotenv
    load_dotenv()
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    run()
