"""
Convert Alvys Factoring Batch exports + Triumph BatchTransactionDetail files
into an Alvys-compatible Upload Payment Report CSV.

The Alvys batch file (DTR_NAME, INVOICE#, INV_DATE, PO, INVAMT) is the
authoritative source for which invoices exist. Triumph data is the lookup for
which of those are funded and what the net amounts are.

Workflow:
  1. In Alvys → Reports → Factoring, open each pending batch and click
     "Batch File" to download the CSV (named XF008_N.csv).
  2. Download Triumph BatchTransactionDetail Excel files from the Triumph portal.
  3. Run this script.
  4. Upload the output CSV(s) to Alvys → Upload Payment Report.

Usage:
  python -m src.triumph_converter \\
      --alvys ~/Downloads/XF008_3.csv [XF008_4.csv ...] \\
      --triumph ~/Downloads/BatchTransactionDetail*.xlsx \\
      [--output ~/Downloads/]
"""

import csv
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

ALVYS_UPLOAD_COLUMNS = [
    "InvoiceNumber",
    "ReferenceNumber",
    "PurchaseDate",
    "CheckNumber",
    "Customer",
    "FeeDays",
    "InvoiceAmount",
    "ActivityType",
    "CheckAmount",
]


# ── Triumph parser ────────────────────────────────────────────────────────────

def _load_triumph_lookup(paths: list[Path]) -> dict[str, dict]:
    """
    Parse all Triumph BatchTransactionDetail Excel files and return a dict
    keyed by invoice number (string) containing only Funded invoices.
    """
    lookup: dict[str, dict] = {}

    for path in paths:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            log.error("Cannot open %s: %s", path.name, e)
            continue

        ws = wb.active
        batch_num = re.sub(r"[^0-9]", "", ws.title) or "UNKNOWN"

        # Find header row
        col: dict[str, int] = {}
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == "Date" and row[1] == "Invoice":
                header_row_idx = i
                col = {name: idx for idx, name in enumerate(row) if name}
                break

        if header_row_idx is None:
            log.warning("No header row in %s — skipping", path.name)
            continue

        count = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if row[1] == "Invoice Total" or row[0] is None:
                break
            if row[col.get("Status", -1)] != "Funded":
                continue

            invoice_num = str(row[col["Invoice"]])
            date_val = row[col["Date"]]
            purchase_date = (
                date_val.strftime("%m/%d/%Y")
                if isinstance(date_val, datetime)
                else str(date_val or "")
            )

            lookup[invoice_num] = {
                "batch_num": batch_num,
                "purchase_date": purchase_date,
                "fee_days": row[col["Age"]],
                "invoice_amount": row[col["Total Amount"]],
                "funded_amount": row[col["Funded Amount"]],
            }
            count += 1

        log.info("Triumph %s → %d funded invoices loaded", path.name, count)

    return lookup


# ── Alvys batch file parser ───────────────────────────────────────────────────

def _load_alvys_batch(path: Path) -> tuple[str, list[dict]]:
    """
    Read an Alvys batch CSV (DTR_NAME, INVOICE#, INV_DATE, PO, INVAMT).
    Returns (batch_label, rows).
    """
    batch_label = path.stem  # e.g. "XF008_3"
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            # Strip trailing empty columns Alvys sometimes appends
            rows.append({k: (v or "").strip() for k, v in row.items() if k})
    return batch_label, rows


# ── Converter ─────────────────────────────────────────────────────────────────

def convert(alvys_path: Path, triumph_lookup: dict, output_dir: Path | None = None) -> Path | None:
    try:
        batch_label, alvys_rows = _load_alvys_batch(alvys_path)
    except Exception as e:
        log.error("Cannot read Alvys batch %s: %s", alvys_path.name, e)
        return None

    output_rows = []
    skipped_unfunded = []

    for row in alvys_rows:
        invoice_num = row.get("INVOICE#", "").strip()
        if not invoice_num:
            continue

        triumph = triumph_lookup.get(invoice_num)
        if not triumph:
            skipped_unfunded.append(invoice_num)
            continue

        output_rows.append({
            "InvoiceNumber": f"T{invoice_num}",
            "ReferenceNumber": row.get("PO", ""),
            "PurchaseDate": triumph["purchase_date"],
            "CheckNumber": triumph["batch_num"],
            "Customer": row.get("DTR_NAME", ""),
            "FeeDays": triumph["fee_days"],
            "InvoiceAmount": triumph["invoice_amount"],
            "ActivityType": "Invoice Purchase",
            "CheckAmount": triumph["funded_amount"],
        })

    if skipped_unfunded:
        log.warning(
            "  %s: %d invoice(s) not yet funded in Triumph — skipped: %s",
            batch_label,
            len(skipped_unfunded),
            ", ".join(skipped_unfunded),
        )

    if not output_rows:
        log.warning("  %s: no funded invoices to upload", batch_label)
        return None

    dest = (output_dir or alvys_path.parent) / f"AlvysPayment_{batch_label}.csv"
    with open(dest, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=ALVYS_UPLOAD_COLUMNS)
        writer.writeheader()
        writer.writerows(output_rows)

    total_inv = sum(float(r["InvoiceAmount"]) for r in output_rows)
    total_funded = sum(float(r["CheckAmount"]) for r in output_rows)

    log.info(
        "  %s → %d/%d invoices · $%s face · $%s funded → %s",
        batch_label,
        len(output_rows),
        len(alvys_rows),
        f"{total_inv:,.2f}",
        f"{total_funded:,.2f}",
        dest.name,
    )
    return dest


# ── CLI ───────────────────────────────────────────────────────────────────────

def main(argv: list[str] | None = None) -> None:
    args = argv if argv is not None else sys.argv[1:]

    alvys_paths: list[Path] = []
    triumph_paths: list[Path] = []
    output_dir: Path | None = None

    i = 0
    while i < len(args):
        if args[i] == "--alvys":
            i += 1
            while i < len(args) and not args[i].startswith("--"):
                alvys_paths.append(Path(args[i]))
                i += 1
        elif args[i] == "--triumph":
            i += 1
            while i < len(args) and not args[i].startswith("--"):
                triumph_paths.append(Path(args[i]))
                i += 1
        elif args[i] == "--output" and i + 1 < len(args):
            output_dir = Path(args[i + 1])
            output_dir.mkdir(parents=True, exist_ok=True)
            i += 2
        else:
            i += 1

    if not alvys_paths or not triumph_paths:
        print(__doc__)
        print("ERROR: --alvys and --triumph are both required.")
        sys.exit(1)

    log.info("Loading Triumph funded invoice data...")
    triumph_lookup = _load_triumph_lookup(triumph_paths)
    log.info("Total funded invoices in lookup: %d", len(triumph_lookup))

    results = [convert(p, triumph_lookup, output_dir) for p in alvys_paths]
    succeeded = [r for r in results if r]
    log.info("Done — %d/%d batch file(s) produced output.", len(succeeded), len(alvys_paths))


if __name__ == "__main__":
    main()
