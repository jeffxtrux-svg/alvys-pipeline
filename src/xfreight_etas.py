"""XFreight ETA report — live X-Trux load tracker, refreshes every 30 min.

Pulls active X-Trux loads from Alvys, current truck GPS from Samsara, and
asks Mapbox Directions API (traffic-aware) for drive time from each truck
to its next undelivered stop. Writes an HTML + Excel snapshot to OneDrive
that overwrites in place, so a single pinned link is always current.

v1 columns (per owner spec, 2026-06-19):
  Shipper | Shipper City | Consignee | Consignee City | Appt | ETA | Delta | Broker

v1 scope:
  - X-Trux entity only
  - Active loads only (Dispatched / In Transit, with an undelivered stop)
  - Trucks with both a matching active Samsara location AND an Alvys load
    are shown; everything else is hidden

Roadmap (deferred — design supports both):
  v2: Customer/broker email notifications when ETA within 45 min of appt
  v2: Contact email + phone column (broker contact if brokered,
      consignee contact if customer-direct)

Env vars (all required unless noted):
  AZURE_TENANT_ID / AZURE_CLIENT_ID / AZURE_CLIENT_SECRET / ONEDRIVE_USER_UPN
  ALVYS_CLIENT_ID / ALVYS_CLIENT_SECRET
  SAMSARA_API_TOKEN
  MAPBOX_TOKEN                       — secret token with directions:read
  TEAMS_ETA_WEBHOOK (optional)       — Power Automate webhook URL; posts an Adaptive
                                       Card when loads go 45+ min late, updates when
                                       the load set changes, and sends an all-clear
                                       card when all loads deliver
  ETA_ONEDRIVE_FOLDER (optional)     — default "ETA"
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv

from src.alvys_client import AlvysClient
from src.samsara_client import SamsaraClient
from src.onedrive_upload import get_token, ensure_folder, upload_file

log = logging.getLogger("xfreight_etas")

CT = ZoneInfo("America/Chicago")
ACTIVE_STATUSES = ["Dispatched", "In Transit"]
_MAPBOX_BASE = "https://api.mapbox.com/directions/v5/mapbox"


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def _entity_is_xtrux(office: str | None) -> bool:
    """Mirrors scorecard_email._entity_group — XFreight + X-Trux office names."""
    if not office:
        return False
    s = str(office).upper()
    return "TRUX" in s or "FREIGHT" in s


def _g(d: dict | None, *path: str, default=None):
    """Walk a nested dict by string path; return default if any hop is missing."""
    cur = d
    for k in path:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
    return cur if cur is not None else default


def _parse_iso(s: str | None) -> datetime | None:
    """Parse Alvys/Samsara ISO timestamp into an aware UTC datetime."""
    if not s:
        return None
    try:
        return datetime.fromisoformat(str(s).replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return None


def _fmt_dt_ct(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.astimezone(CT).strftime("%a %b %d, %I:%M%p").replace("AM", "am").replace("PM", "pm")


def _fmt_date_local(dt: datetime | None) -> str:
    """Date only (no time) in the timestamp's OWN zone, e.g. 'Mon Jun 30'.

    For date-only appointments: the receiver gave a calendar date in their local
    zone, so converting to Central could flip the day (an Eastern midnight is the
    prior evening in CT). Format the date as written instead of converting.
    """
    if dt is None:
        return ""
    return dt.strftime("%a %b %d")


def _is_midnight(iso: str | None) -> bool:
    """True when an ISO timestamp falls exactly on midnight in its own offset
    (i.e. it carries a date but no time-of-day). Alvys emits date-only stop
    windows as Begin==End==00:00:00 in the stop's local zone."""
    dt = _parse_iso(iso)
    return dt is not None and (dt.hour, dt.minute, dt.second) == (0, 0, 0)


def _fmt_delta(minutes: int | None) -> tuple[str, str]:
    """Return (display_text, css_color) for the Delta column."""
    if minutes is None:
        return ("—", "#999")
    if minutes <= _LATE_THRESHOLD_MIN:
        return (f"{-minutes} min late", RED)
    if minutes < 0:
        return (f"{-minutes} min late", AMBER)
    if minutes <= 30:
        return (f"{minutes} min early", "#16a34a")  # green
    return (f"{minutes} min early", INK)


def _fmt_appt_age(appt_dt: datetime | None, now: datetime) -> str:
    """Human 'how long ago' for a past appointment, e.g. '6d ago' / '18h ago'."""
    if appt_dt is None:
        return ""
    hrs = (now - appt_dt).total_seconds() / 3600
    if hrs < 0:
        return ""
    if hrs >= 48:
        return f"{int(round(hrs / 24))}d ago"
    return f"{int(round(hrs))}h ago"


def _is_appt_stale(appt_dt: datetime | None, now: datetime) -> bool:
    """True when an appointment is more than _STALE_APPT_HOURS in the past.

    A stale appt means the truck is almost certainly already delivered (with
    ArrivedAt unset in Alvys) or the appt was never rescheduled — so the precise
    "Xh late" figure is a data gap, not a live tracking event.
    """
    return bool(appt_dt) and appt_dt < now - timedelta(hours=_STALE_APPT_HOURS)


# ----------------------------------------------------------------------
# HOS helpers
# ----------------------------------------------------------------------
_HOS_REST_SECONDS = 10 * 3600  # federal property-carrier mandatory break


def _norm_name(name: str | None) -> str:
    """Normalize a person name for matching: lowercase, drop punctuation,
    digits and common suffixes, collapse whitespace. 'John A. Smith Jr.' →
    'john a smith'."""
    s = (name or "").lower()
    s = re.sub(r"[^a-z\s]", " ", s)                      # drop punctuation/digits
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", " ", s)        # drop generational suffixes
    return re.sub(r"\s+", " ", s).strip()


def _name_keys(name: str | None) -> list[str]:
    """Match keys for a name, most specific first: full normalized name, then
    first+last (drops middle names/initials so 'John A Smith' matches 'John
    Smith'). Deliberately never falls back to last-name-only — that would risk
    matching the wrong driver and a wrong HOS clock is worse than none."""
    n = _norm_name(name)
    if not n:
        return []
    keys = [n]
    parts = n.split()
    if len(parts) >= 2:
        fl = f"{parts[0]} {parts[-1]}"
        if fl != n:
            keys.append(fl)
    return keys


def _build_hos_index(hos_clocks: list[dict]) -> dict[str, int]:
    """{match_key: remaining_drive_seconds} from Samsara HOS clocks.

    Keyed by the full normalized driver name AND, when unambiguous, a
    first+last alias — so an Alvys name carrying a middle name/initial still
    matches the Samsara clock. Ambiguous first+last aliases (two drivers who
    share them) are dropped rather than risk a wrong match.
    """
    by_full: dict[str, int] = {}
    fl_owners: dict[str, set] = {}   # first_last -> set of distinct full names
    fl_value: dict[str, int] = {}
    for rec in hos_clocks:
        drv = rec.get("driver") or {}
        name = (drv.get("name") or "").strip()
        if not name:
            continue
        drive = (rec.get("clocks") or {}).get("drive") or {}
        # Samsara HOS endpoint uses driveRemainingDurationMs (confirmed from live data)
        remaining_ms = (drive.get("driveRemainingDurationMs")
                        or drive.get("remainingMs"))
        remaining_s = drive.get("remainingSeconds")
        if remaining_ms is not None:
            val = int(remaining_ms) // 1000
        elif remaining_s is not None:
            val = int(remaining_s)
        else:
            continue
        full = _norm_name(name)
        if not full:
            continue
        by_full[full] = val
        parts = full.split()
        if len(parts) >= 2:
            fl = f"{parts[0]} {parts[-1]}"
            fl_owners.setdefault(fl, set()).add(full)
            fl_value[fl] = val
    idx = dict(by_full)
    for fl, owners in fl_owners.items():
        if len(owners) == 1 and fl not in idx:   # unambiguous, doesn't shadow a full name
            idx[fl] = fl_value[fl]
    return idx


def _hos_remaining(hos_index: dict[str, int], driver_name: str) -> int | None:
    """Return remaining drive seconds for a driver, or None if not found.
    Tries the full normalized name, then an unambiguous first+last alias."""
    if not driver_name or not hos_index:
        return None
    for key in _name_keys(driver_name):
        if key in hos_index:
            return hos_index[key]
    return None


def _fmt_hos(remaining_seconds: int | None) -> str:
    """Format remaining drive seconds as 'Xh Ym' or '—'."""
    if remaining_seconds is None:
        return "—"
    if remaining_seconds <= 0:
        return "0h 0m"
    h, m = divmod(remaining_seconds // 60, 60)
    return f"{h}h {m}m"


# ----------------------------------------------------------------------
# Alvys load extraction
# ----------------------------------------------------------------------
def _next_undelivered_stop(load: dict) -> dict | None:
    """Return the next stop a truck still has to hit (no ArrivedAt)."""
    stops = load.get("Stops") or []
    for stop in stops:
        if not stop.get("ArrivedAt"):
            return stop
    return None


def _remaining_undelivered_stops(load: dict) -> list[dict]:
    """All stops not yet arrived at, in order."""
    return [s for s in (load.get("Stops") or []) if not s.get("ArrivedAt")]


def _is_date_only_window(begin: str | None, end: str | None) -> bool:
    """True when a stop window carries a date but no real time-of-day.

    Alvys returns these as Begin==End at midnight (also seen as End-only at
    midnight). There is no hard time deadline — the receiver gave a date — so
    tracking lateness against the midnight boundary would make a same-day
    daytime delivery look hours "late".
    """
    if not end or not _is_midnight(end):
        return False
    # End is midnight → date-only, unless Begin supplies a real daytime time.
    return (not begin) or (begin == end) or _is_midnight(begin)


def _stop_appt_iso(stop: dict) -> str | None:
    """ISO string used for delta calculation (ETA vs. deadline).
    - APPT  → AppointmentDate (fixed time; truck is late if ETA > appt)
    - WINDOW → StopWindow.End (late only when truck misses the window close)
    - FCFS with End → StopWindow.End (receiver closes at End; arriving later = late)
    - FCFS with no End → None (open-ended FCFS; no hard deadline to track)
    - Date-only window (Begin==End at midnight) → None (no hard time deadline)
    """
    stype = (stop.get("ScheduleType") or "").upper()
    window = stop.get("StopWindow") or {}
    if stype == "APPT":
        return stop.get("AppointmentDate")
    end = window.get("End")
    begin = window.get("Begin")
    # Date-only window — a date with no time-of-day carries no hard deadline.
    if _is_date_only_window(begin, end):
        return None
    if end:
        # Both WINDOW and FCFS: if a window closes, arriving after close = late.
        return end
    if stype == "FCFS":
        # FCFS with no End — truly open-ended (dock open, first come first served).
        return None
    # WINDOW or unknown type with only Begin (or AppointmentDate fallback)
    return begin or stop.get("AppointmentDate")


def _stop_window_begin_iso(stop: dict) -> str | None:
    """Window open time for display purposes (shown as 'Begin – End' in Appt column).
    Only returned when End also exists AND the window is a real time range — never
    for a date-only window (would render a degenerate '12:00am – 12:00am') or a
    single fixed time (Begin==End)."""
    stype = (stop.get("ScheduleType") or "").upper()
    if stype == "APPT":
        return None
    window = stop.get("StopWindow") or {}
    begin, end = window.get("Begin"), window.get("End")
    if _is_date_only_window(begin, end):
        return None
    # Real range only: End present and Begin is a distinct earlier time.
    if end and begin and begin != end:
        return begin
    return None


def _stop_date_only_iso(stop: dict) -> str | None:
    """For a date-only window, the calendar date to DISPLAY (no time deadline)."""
    window = stop.get("StopWindow") or {}
    begin, end = window.get("Begin"), window.get("End")
    if _is_date_only_window(begin, end):
        return end or begin
    return None


def _is_brokered(load: dict) -> bool:
    return str(load.get("BrokerageStatus") or "").lower() == "brokered"


def _extract_load_row(load: dict, trucks_by_id: dict, trips_by_load: dict,
                      drivers_by_id: dict | None = None,
                      users_by_id: dict | None = None) -> dict | None:
    """Pull the v1 report columns out of one Alvys load record. Returns None
    if the load isn't routable (no truck assignment, no undelivered stop,
    or no geocoded destination).

    Truck assignment lives on the Trip (not the Load). We join by LoadNumber.
    Coordinates are on the Stop under a 'Coordinates' key (not Address.Latitude).
    """
    load_num = str(load.get("LoadNumber") or "")
    trip = trips_by_load.get(load_num) if load_num else None

    truck_name = None
    driver_name = None
    if trip:
        truck_obj = trip.get("Truck") or {}
        if isinstance(truck_obj, dict):
            truck_name = (truck_obj.get("TruckNum") or truck_obj.get("TruckNumber")
                         or truck_obj.get("Number") or truck_obj.get("Name"))
            if not truck_name:
                truck_id = truck_obj.get("Id")
                if truck_id:
                    truck_name = trucks_by_id.get(str(truck_id)) or None

        driver1_obj = trip.get("Driver1") or {}
        if isinstance(driver1_obj, dict):
            driver_name = (driver1_obj.get("FullName") or driver1_obj.get("Name")
                           or driver1_obj.get("DisplayName"))
            if not driver_name and drivers_by_id:
                driver_id = driver1_obj.get("Id")
                if driver_id:
                    driver_name = drivers_by_id.get(str(driver_id))

    if not truck_name:
        return None

    remaining_stops = _remaining_undelivered_stops(load)
    if not remaining_stops:
        return None

    # Build ordered waypoints for every remaining stop that has valid coordinates.
    # Coordinates live under the "Coordinates" key, not inside Address.
    route_waypoints: list[dict] = []
    for stop in remaining_stops:
        coords = stop.get("Coordinates") or {}
        lat = coords.get("Latitude") if isinstance(coords, dict) else None
        lng = coords.get("Longitude") if isinstance(coords, dict) else None
        if lat is not None and lng is not None:
            route_waypoints.append({"lat": float(lat), "lng": float(lng)})

    if not route_waypoints:
        return None

    stops = load.get("Stops") or []
    first_stop = stops[0] if stops else {}
    last_stop = stops[-1] if stops else {}

    _stype = (last_stop.get("ScheduleType") or "").upper() or "—"
    _win = last_stop.get("StopWindow") or {}
    log.info("load %s last_stop sched: type=%s appt=%s window_begin=%s window_end=%s",
             load.get("LoadNumber"), _stype,
             last_stop.get("AppointmentDate"), _win.get("Begin"), _win.get("End"))

    return {
        "load_no": load.get("LoadNumber") or load.get("Number"),
        "truck_name": str(truck_name),
        "shipper": _g(first_stop, "CompanyName") or _g(first_stop, "Address", "Street") or "",
        "shipper_city": _g(first_stop, "Address", "City") or "",
        "shipper_state": _g(first_stop, "Address", "State") or "",
        "consignee": _g(last_stop, "CompanyName") or _g(last_stop, "Address", "Street") or "",
        "consignee_city": _g(last_stop, "Address", "City") or "",
        "consignee_state": _g(last_stop, "Address", "State") or "",
        "appt_dt": _parse_iso(_stop_appt_iso(last_stop)),
        "appt_window_begin_dt": _parse_iso(_stop_window_begin_iso(last_stop)),
        # Date-only window (date, no time deadline): the date to display.
        "_date_only_dt": _parse_iso(_stop_date_only_iso(last_stop)),
        "appt_stype": _stype,          # "APPT" / "WINDOW" / "FCFS" / "—"
        # For FCFS-with-Begin-only: the dock open time, shown as "FCFS 8:00am".
        # Distinct from appt_window_begin_dt (which is the left side of "Begin – End").
        "_fcfs_open_dt": (
            _parse_iso(_win.get("Begin"))
            if _stype == "FCFS" and not _win.get("End") and _win.get("Begin")
            else None
        ),
        "route_waypoints": route_waypoints,
        "broker": load.get("CustomerName") if _is_brokered(load) else "",
        "customer_name": load.get("CustomerName") or "",
        "office": _g(load, "Office", "Name") or _g(load, "Trip", "Office", "Name") or "",
        "driver_name": driver_name or "",
        "sales_agent": (users_by_id or {}).get(
            str(load.get("CustomerSalesAgentId") or "")) or "",
    }


# ----------------------------------------------------------------------
# Samsara location join
# ----------------------------------------------------------------------
def _locations_by_truck_name(samsara_locations: list[dict]) -> dict:
    """{truck_name: {lat, lng, ts}} from Samsara /fleet/vehicles/locations.

    Also adds an unambiguous digits-only alias per vehicle (so an Alvys truck
    number '42187' matches a Samsara name like 'Truck 42187' or '#42187').
    Aliases never shadow an exact name and are dropped when two vehicles share
    the same digits.
    """
    out: dict[str, dict] = {}
    digit_owners: dict[str, set] = {}
    digit_val: dict[str, dict] = {}
    for rec in samsara_locations:
        name = rec.get("name") or rec.get("vehicle", {}).get("name")
        loc = rec.get("location") or {}
        lat = loc.get("latitude")
        lng = loc.get("longitude")
        ts = _parse_iso(loc.get("time") or rec.get("time"))
        if name and lat is not None and lng is not None:
            key = str(name).strip()
            val = {"lat": float(lat), "lng": float(lng), "ts": ts}
            out[key] = val
            digits = re.sub(r"\D", "", key)
            if digits:
                digit_owners.setdefault(digits, set()).add(key)
                digit_val[digits] = val
    for digits, owners in digit_owners.items():
        if len(owners) == 1 and digits not in out:
            out[digits] = digit_val[digits]
    return out


def _lookup_truck_gps(locs_by_truck: dict, truck_name) -> dict | None:
    """Find a truck's GPS by exact name, then by its digits-only alias."""
    if not truck_name:
        return None
    hit = locs_by_truck.get(str(truck_name).strip())
    if hit:
        return hit
    digits = re.sub(r"\D", "", str(truck_name))
    return locs_by_truck.get(digits) if digits else None


# ----------------------------------------------------------------------
# Mapbox routing
# ----------------------------------------------------------------------
def _mapbox_duration_seconds(
    token: str, from_lat: float, from_lng: float,
    waypoints: list[dict],  # [{lat, lng}, ...] — all remaining stops in order
    timeout: int = 15,
) -> float | None:
    """Query Mapbox Directions and return total drive-time in seconds for the
    full remaining route: truck → stop1 → stop2 → … → final stop.

    Tries driving-traffic (requires sk. token) first; falls back to driving
    (works with pk. token) on 401 so the report works immediately with a
    public token and upgrades automatically when rotated to secret token.
    """
    coords = ";".join(
        [f"{from_lng},{from_lat}"] + [f"{w['lng']},{w['lat']}" for w in waypoints]
    )
    params = {"access_token": token, "geometries": "geojson", "overview": "false"}
    for profile in ("driving-traffic", "driving"):
        url = f"{_MAPBOX_BASE}/{profile}/{coords}"
        try:
            resp = requests.get(url, params=params, timeout=timeout)
            if resp.status_code == 401 and profile == "driving-traffic":
                log.info("Mapbox driving-traffic → 401 (pk. token?), retrying with driving")
                continue
            if resp.status_code != 200:
                log.warning("Mapbox %s→%s HTTP %d: %s",
                            profile, coords[:30], resp.status_code, resp.text[:200])
                return None
            routes = (resp.json() or {}).get("routes") or []
            return float(routes[0].get("duration") or 0) if routes else None
        except Exception as e:
            log.warning("Mapbox %s request failed: %s", profile, e)
            return None
    return None


# ----------------------------------------------------------------------
# Teams alert
# ----------------------------------------------------------------------
_LATE_THRESHOLD_MIN = -45
_TEAMS_STATE_FILE = "eta_state.json"
_ETA_LOG_FILE = "eta_log.csv"
_ETA_LOG_KEEP_DAYS = 90   # rows older than this are pruned on each run
_GPS_STALE_MINUTES = 45  # GPS fix older than this is treated as unreliable
# An appointment more than this many hours in the past is treated as stale: the
# truck is almost certainly already delivered with ArrivedAt unset in Alvys, or
# the appt was never rescheduled after the load slipped. Either way the precise
# "Xh late" figure is misleading, so these are flagged "verify delivery" instead
# of firing a hard late alert (a delta of −45 min on a same-day appt is real;
# −9000 min on a 6-day-old appt is a data gap, not a live tracking event).
_STALE_APPT_HOURS = 24


# ----------------------------------------------------------------------
# Teams alert — shared card builder
# ----------------------------------------------------------------------
def _build_teams_card(late_rows: list[dict]) -> dict:
    """Build the Adaptive Card dict for the current set of late loads."""
    body: list[dict] = [
        {"type": "TextBlock", "text": "⚠️ Drivers Running Late",
         "weight": "Bolder", "size": "Large", "color": "Attention", "wrap": True},
        {"type": "TextBlock",
         "text": f"{len(late_rows)} load(s) are 45+ min behind schedule — as of "
                 f"{datetime.now(CT):%I:%M %p CT}",
         "size": "Small", "spacing": "None", "wrap": True},
    ]
    for idx, r in enumerate(sorted(late_rows, key=lambda x: x.get("delta_min") or 0)):
        mins_late = abs(r["delta_min"])
        hrs, m = divmod(mins_late, 60)
        late_str = f"{hrs}h {m}m late" if hrs else f"{m}m late"
        dest = (f"{r['consignee_city']}, {r['consignee_state']}".strip(", ")
                or r["consignee"] or "—")
        driver = r.get("driver_name") or "—"
        p = f"l{idx}"
        body.append({
            "type": "Container",
            "separator": True,
            "spacing": "Medium",
            "items": [
                {
                    "type": "ColumnSet",
                    "columns": [
                        {
                            "type": "Column", "width": "stretch",
                            "items": [{"type": "TextBlock",
                                       "text": f"**Truck {r['truck_name']}** — {driver}",
                                       "wrap": True}],
                        },
                        {
                            "type": "Column", "width": "auto",
                            "items": [{"type": "TextBlock", "text": f"**{late_str}**",
                                       "color": "Attention", "weight": "Bolder",
                                       "wrap": False}],
                        },
                    ],
                },
                {
                    "type": "FactSet",
                    "spacing": "Small",
                    "facts": [
                        {"title": "Load #", "value": str(r.get("load_no") or "—")},
                        {"title": "Broker" if r.get("broker") else "Customer",
                         "value": r.get("customer_name") or "—"},
                        {"title": "Destination", "value": dest},
                        {"title": "Appt",
                         "value": (
                             f"{_fmt_dt_ct(r['appt_window_begin_dt'])} – {_fmt_dt_ct(r['appt_dt'])}"
                             if r.get("appt_window_begin_dt")
                             else (f"FCFS {_fmt_dt_ct(r['_fcfs_open_dt'])}"
                                   if r.get("_fcfs_open_dt")
                                   else _fmt_dt_ct(r.get("appt_dt")) or "—")
                         )},
                        {"title": "ETA", "value": _fmt_dt_ct(r.get("eta_dt")) or "—"},
                        {"title": "HOS Left",
                         "value": (_fmt_hos(r.get("hos_remaining_s"))
                                   + (" ⚠️ +10h rest factored in" if r.get("hos_delay") else ""))},
                        {"title": "Sales Agent", "value": r.get("sales_agent") or "—"},
                    ],
                },
                {
                    "type": "ActionSet",
                    "spacing": "Small",
                    "actions": [
                        {"type": "Action.ToggleVisibility", "title": "✓ Notified Customer",
                         "targetElements": [f"{p}_cust"]},
                        {"type": "Action.ToggleVisibility", "title": "✓ Notified Broker",
                         "targetElements": [f"{p}_brkr"]},
                        {"type": "Action.ToggleVisibility", "title": "✓ Notified Shipper",
                         "targetElements": [f"{p}_shpr"]},
                        {"type": "Action.ToggleVisibility", "title": "📅 Rescheduled",
                         "targetElements": [f"{p}_rsch"]},
                        {"type": "Action.ToggleVisibility", "title": "🔧 Working On It",
                         "targetElements": [f"{p}_work"]},
                    ],
                },
                {"type": "TextBlock", "id": f"{p}_cust", "isVisible": False,
                 "text": "✅ Customer notified", "color": "Good",
                 "size": "Small", "spacing": "None", "wrap": False},
                {"type": "TextBlock", "id": f"{p}_brkr", "isVisible": False,
                 "text": "✅ Broker notified", "color": "Good",
                 "size": "Small", "spacing": "None", "wrap": False},
                {"type": "TextBlock", "id": f"{p}_shpr", "isVisible": False,
                 "text": "✅ Shipper notified", "color": "Good",
                 "size": "Small", "spacing": "None", "wrap": False},
                {"type": "TextBlock", "id": f"{p}_rsch", "isVisible": False,
                 "text": "📅 Rescheduled", "color": "Good",
                 "size": "Small", "spacing": "None", "wrap": False},
                {"type": "TextBlock", "id": f"{p}_work", "isVisible": False,
                 "text": "🔧 Working on it", "color": "Warning",
                 "size": "Small", "spacing": "None", "wrap": False},
            ],
        })
    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": body,
        "msteams": {"width": "Full"},
    }


# ----------------------------------------------------------------------
# Teams alert — OneDrive state helpers
# ----------------------------------------------------------------------
_GRAPH_BASE = "https://graph.microsoft.com/v1.0"


def _load_teams_state(token: str, user_upn: str, folder: str) -> dict:
    """Download eta_state.json from OneDrive; return {} if absent or unreadable."""
    path = f"{folder}/{_TEAMS_STATE_FILE}"
    url = f"{_GRAPH_BASE}/users/{user_upn}/drive/root:/{path}:/content"
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=15)
        if resp.status_code == 200:
            return resp.json()
        if resp.status_code != 404:
            log.warning("Teams state load HTTP %d", resp.status_code)
    except Exception as exc:
        log.warning("Teams state load failed: %s", exc)
    return {}


def _save_teams_state(token: str, user_upn: str, folder: str, state: dict) -> None:
    """Upload eta_state.json to OneDrive (simple PUT, small file)."""
    path = f"{folder}/{_TEAMS_STATE_FILE}"
    url = f"{_GRAPH_BASE}/users/{user_upn}/drive/root:/{path}:/content"
    try:
        resp = requests.put(
            url,
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            data=json.dumps(state, default=str).encode(),
            timeout=15,
        )
        if resp.status_code not in (200, 201):
            log.warning("Teams state save HTTP %d: %s",
                        resp.status_code, resp.text[:200])
    except Exception as exc:
        log.warning("Teams state save failed: %s", exc)


# ----------------------------------------------------------------------
# Teams alert — state-tracked webhook (posts only on change; all-clear on resolve)
# ----------------------------------------------------------------------
def _build_clear_card() -> dict:
    """Adaptive Card posted when all previously-late loads have cleared."""
    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": [
            {"type": "TextBlock", "text": "✅ All Loads Back On Schedule",
             "weight": "Bolder", "size": "Large", "color": "Good", "wrap": True},
            {"type": "TextBlock",
             "text": f"No loads are 45+ min behind schedule — as of "
                     f"{datetime.now(CT):%I:%M %p CT}",
             "size": "Small", "spacing": "None", "wrap": True},
        ],
        "msteams": {"width": "Full"},
    }


def _build_resolved_card(resolved_load_nos: set[str]) -> dict:
    """Adaptive Card posted for each batch of loads that drop off the late list."""
    loads_str = ", ".join(f"#{ln}" for ln in sorted(resolved_load_nos))
    return {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.4",
        "body": [
            {"type": "TextBlock",
             "text": f"✅ Resolved — Load{'s' if len(resolved_load_nos) > 1 else ''} "
                     f"{loads_str} no longer behind schedule",
             "weight": "Bolder", "size": "Medium", "color": "Good", "wrap": True},
            {"type": "TextBlock",
             "text": f"As of {datetime.now(CT):%I:%M %p CT} — "
                     "previous alert for this load is superseded.",
             "size": "Small", "color": "Good", "spacing": "None", "wrap": True},
        ],
        "msteams": {"width": "Full"},
    }


def _sync_teams_webhook(webhook_url: str, token: str, user_upn: str, folder: str,
                        late_rows: list[dict]) -> None:
    """Post to the Teams webhook when the set of late loads changes OR an
    appointment is updated for a load that is already in the alerted set.

    - New load(s) became 45+ min late → POST card with all current late loads.
    - Load(s) resolved → POST "✅ Resolved" card for those loads THEN updated list.
    - All late loads cleared → POST "✅ All loads back on schedule" card.
    - Appointment changed for an already-alerted load → POST updated card.
    - Same late-load set + same appointments → skip (no card posted).

    State is stored in OneDrive (eta_state.json) and persists across 30-min runs.
    """
    state = _load_teams_state(token, user_upn, folder)
    prev_load_nos = set(state.get("alerted_load_nos") or [])
    prev_appts: dict = state.get("alerted_appts") or {}

    curr_load_nos = {str(r["load_no"]) for r in late_rows if r.get("load_no")}
    curr_appts = {
        str(r["load_no"]): r["appt_dt"].isoformat() if r.get("appt_dt") else ""
        for r in late_rows if r.get("load_no")
    }

    # Detect appointment changes for loads that are in both sets
    appt_changed_loads = [
        ln for ln in curr_load_nos & prev_load_nos
        if curr_appts.get(ln) != prev_appts.get(ln)
    ]

    if curr_load_nos == prev_load_nos and not appt_changed_loads:
        log.info("Teams: late-load set unchanged (%d loads) — no card posted",
                 len(curr_load_nos))
        return

    if appt_changed_loads:
        log.info("Teams: appointment changed for load(s) %s — posting updated card",
                 appt_changed_loads)

    resolved_loads = prev_load_nos - curr_load_nos

    def _post(card: dict) -> bool:
        payload = {
            "type": "message",
            "attachments": [{"contentType": "application/vnd.microsoft.card.adaptive",
                             "content": card}],
        }
        try:
            resp = requests.post(webhook_url, json=payload, timeout=15)
            if resp.status_code not in (200, 202):
                log.warning("Teams webhook HTTP %d: %s",
                            resp.status_code, resp.text[:300])
                return False
            return True
        except Exception as exc:
            log.warning("Teams webhook failed: %s", exc)
            return False

    # Post a "resolved" card for any load(s) that just dropped off the late list.
    # This explicitly supersedes the previous alert so the channel history is clear.
    if resolved_loads:
        log.info("Teams: posting resolved card for load(s) %s", sorted(resolved_loads))
        _post(_build_resolved_card(resolved_loads))

    if curr_load_nos:
        card = _build_teams_card(late_rows)
        new_state: dict = {
            "alerted_load_nos": sorted(curr_load_nos),
            "alerted_appts": curr_appts,
            "last_alerted": datetime.now(timezone.utc).isoformat(),
        }
        log_msg = "Teams: posted updated alert for %d late load(s)"
    else:
        card = _build_clear_card()
        new_state = {}
        log_msg = "Teams: posted all-clear notification (%d loads cleared)"

    if not _post(card):
        return
    log.info(log_msg, len(prev_load_nos - curr_load_nos) if not curr_load_nos
             else len(curr_load_nos))
    _save_teams_state(token, user_upn, folder, new_state)


# ----------------------------------------------------------------------
# Historical ETA log
# ----------------------------------------------------------------------
def _append_eta_log(token: str, user_upn: str, folder: str,
                    run_ts: datetime,
                    rows_with_eta: list[dict],
                    untracked: list[dict]) -> None:
    """Append this run's load snapshots to a rolling CSV on OneDrive.

    One row per load per run — both tracked (ETA computed) and untracked
    (ETA blank, reason filled). Rows older than _ETA_LOG_KEEP_DAYS are
    pruned on each write so the file stays bounded.

    Columns: run_ts, load_no, truck, driver, consignee, consignee_city,
             consignee_state, appt_iso, eta_iso, delta_min, late_45plus,
             gps_stale, appt_stale, untracked, reason
    """
    import csv as _csv
    import io as _io

    _FIELDS = [
        "run_ts", "load_no", "truck", "driver", "consignee",
        "consignee_city", "consignee_state", "appt_iso", "eta_iso",
        "delta_min", "late_45plus", "gps_stale", "appt_stale", "untracked", "reason",
    ]

    path = f"{folder}/{_ETA_LOG_FILE}"
    url = f"{_GRAPH_BASE}/users/{user_upn}/drive/root:/{path}:/content"
    cutoff = run_ts - timedelta(days=_ETA_LOG_KEEP_DAYS)

    # Download and prune existing log
    existing: list[dict] = []
    try:
        resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=20)
        if resp.status_code == 200:
            reader = _csv.DictReader(_io.StringIO(resp.text))
            for row in reader:
                try:
                    if datetime.fromisoformat(row["run_ts"]) >= cutoff:
                        existing.append(row)
                except (KeyError, ValueError):
                    pass
            log.info("ETA log: loaded %d existing rows (after 90d prune)", len(existing))
        elif resp.status_code != 404:
            log.warning("ETA log download HTTP %d — starting fresh", resp.status_code)
    except Exception as exc:
        log.warning("ETA log download failed: %s — starting fresh", exc)

    # Build new rows for this run
    run_ts_iso = run_ts.isoformat()
    new_rows: list[dict] = []

    for r in rows_with_eta:
        delta = r.get("delta_min")
        new_rows.append({
            "run_ts": run_ts_iso,
            "load_no": str(r.get("load_no") or ""),
            "truck": str(r.get("truck_name") or ""),
            "driver": str(r.get("driver_name") or ""),
            "consignee": str(r.get("consignee") or ""),
            "consignee_city": str(r.get("consignee_city") or ""),
            "consignee_state": str(r.get("consignee_state") or ""),
            "appt_iso": r["appt_dt"].isoformat() if r.get("appt_dt") else "",
            "eta_iso": r["eta_dt"].isoformat() if r.get("eta_dt") else "",
            "delta_min": str(delta) if delta is not None else "",
            "late_45plus": "1" if (delta is not None and delta <= _LATE_THRESHOLD_MIN) else "0",
            "gps_stale": "1" if r.get("gps_stale") else "0",
            "appt_stale": "1" if r.get("appt_stale") else "0",
            "untracked": "0",
            "reason": "",
        })

    for r in untracked:
        new_rows.append({
            "run_ts": run_ts_iso,
            "load_no": str(r.get("load_no") or ""),
            "truck": str(r.get("truck") or ""),
            "driver": str(r.get("driver") or ""),
            "consignee": str(r.get("consignee") or ""),
            "consignee_city": str(r.get("consignee_city") or ""),
            "consignee_state": str(r.get("consignee_state") or ""),
            "appt_iso": r["appt_dt"].isoformat() if r.get("appt_dt") else "",
            "eta_iso": "",
            "delta_min": "",
            "late_45plus": "",
            "gps_stale": "",
            "appt_stale": "",
            "untracked": "1",
            "reason": str(r.get("reason") or ""),
        })

    all_rows = existing + new_rows

    # Serialize and upload
    buf = _io.StringIO()
    writer = _csv.DictWriter(buf, fieldnames=_FIELDS, lineterminator="\r\n",
                             extrasaction="ignore")
    writer.writeheader()
    writer.writerows(all_rows)
    csv_bytes = buf.getvalue().encode("utf-8")

    try:
        resp = requests.put(
            url,
            headers={"Authorization": f"Bearer {token}", "Content-Type": "text/csv"},
            data=csv_bytes,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            log.info("ETA log: %d total rows (%d new) → %s", len(all_rows), len(new_rows), path)
        else:
            log.warning("ETA log upload HTTP %d: %s", resp.status_code, resp.text[:200])
    except Exception as exc:
        log.warning("ETA log upload failed: %s", exc)


# ----------------------------------------------------------------------
# Report rendering
# ----------------------------------------------------------------------
INK = "#1a1a1a"


def _fmt_appt_cell(row: dict) -> str:
    """Format the Appt column.
    - APPT          → fixed time   e.g. "Tue Jun 23, 08:00am"
    - WINDOW/FCFS with End → "Begin – End"  e.g. "8:00am – 5:00pm"
    - FCFS with Begin only (no End) → "FCFS 8:00am"  (dock open time, no close deadline)
    - Date-only window → "Mon Jun 30 (any time)"  (a date, no hard time deadline)
    - No schedule info → "—"
    """
    end_str = _fmt_dt_ct(row.get("appt_dt"))
    begin_dt = row.get("appt_window_begin_dt")
    if begin_dt and end_str:
        return f"{_fmt_dt_ct(begin_dt)}&nbsp;–&nbsp;{end_str}"
    if end_str:
        return end_str
    # FCFS with Begin only — show open time with label so dispatch knows the dock opens then
    stype = row.get("appt_stype", "")
    if stype == "FCFS":
        win_begin = row.get("_fcfs_open_dt")
        if win_begin:
            return f"FCFS&nbsp;{_fmt_dt_ct(win_begin)}"
    # Date-only window — show the date and make clear there is no time deadline.
    date_only = row.get("_date_only_dt")
    if date_only:
        return (f"{_fmt_date_local(date_only)}&nbsp;"
                f"<span style='color:{MUTE};font-size:10px;'>(any time)</span>")
    return "—"
MUTE = "#6b6b6b"
AMBER = "#d97706"
LINE = "#e5e5e5"
RED = "#c41e2a"
TILEBG = "#fafafa"
FONT = ("font-family:-apple-system,'Helvetica Neue',Helvetica,Arial,sans-serif;"
        "font-size:13px;color:#1a1a1a;")


def _render_html(rows: list[dict], generated_at: datetime,
                 untracked: list[dict] | None = None) -> str:
    untracked = untracked or []
    # Live-late count EXCLUDES stale appointments — a 6-day-old appt isn't a live
    # late event (see _STALE_APPT_HOURS), so it must not inflate the red banner.
    late_count = sum(1 for r in rows
                     if (r.get("delta_min") or 0) <= _LATE_THRESHOLD_MIN
                     and not r.get("appt_stale"))
    stale_count = sum(1 for r in rows if r.get("appt_stale"))

    if late_count:
        strip_bg = RED
        strip_msg = f"⚠ {late_count} load{'s' if late_count != 1 else ''} 45+ min late"
    elif stale_count and rows:
        strip_bg = AMBER
        strip_msg = (f"{stale_count} load{'s' if stale_count != 1 else ''} with a stale "
                     f"appointment — verify delivery")
    elif rows:
        strip_bg = "#16a34a"
        strip_msg = "All loads on schedule"
    else:
        strip_bg = MUTE
        strip_msg = "No active X-Trux loads"

    untracked_badge = (f" &middot; {len(untracked)} untracked" if untracked else "")
    # Show the stale chip on the sub-line unless the strip already leads with it.
    stale_badge = (f" &middot; {stale_count} stale appt"
                   if (stale_count and late_count) else "")

    if not rows:
        body = (f"<div style='padding:40px;text-align:center;color:{MUTE};font-size:14px;'>"
                f"No active X-Trux loads to display.</div>")
    else:
        # Sort: latest (worst delta first), then by appt time
        rows = sorted(
            rows,
            key=lambda r: ((r.get("delta_min") if r.get("delta_min") is not None else 9999),
                           r.get("appt_dt") or datetime.max.replace(tzinfo=timezone.utc)),
        )
        thead = (
            f"<thead><tr style='background:{TILEBG};border-bottom:2px solid {INK};'>"
            + "".join(
                f"<th style='padding:8px 10px;text-align:left;font-size:10px;"
                f"text-transform:uppercase;letter-spacing:0.8px;color:{MUTE};'>"
                f"{h}</th>"
                for h in ("Truck", "Driver", "Shipper", "Shipper City", "Consignee",
                          "Consignee City", "Appt", "ETA", "Delta", "HOS Left", "Broker"))
            + "</tr></thead>"
        )

        tbody_rows = ""
        _now_ct = generated_at
        for r in rows:
            delta_txt, delta_color = _fmt_delta(r.get("delta_min"))
            shipper_loc = f"{r['shipper_city']}, {r['shipper_state']}".strip(", ")
            consignee_loc = f"{r['consignee_city']}, {r['consignee_state']}".strip(", ")
            hos_txt = _fmt_hos(r.get("hos_remaining_s"))
            hos_color = RED if r.get("hos_delay") else MUTE

            # ETA cell: prefix ~ and show GPS age if fix is stale
            eta_str = _fmt_dt_ct(r.get("eta_dt"))
            if r.get("gps_stale") and r.get("gps_age_min") is not None:
                eta_str = (f"~{eta_str}&nbsp;"
                           f"<span style='color:{AMBER};font-size:10px;font-weight:400;'>"
                           f"&#9888; GPS {r['gps_age_min']}m old</span>")
                delta_txt = f"~{delta_txt}"

            # Stale appointment: don't show a precise (and misleading) "Xh late".
            # The appt is days old → the load is almost certainly already
            # delivered or never rescheduled. Flag for a status check instead.
            if r.get("appt_stale"):
                _age = _fmt_appt_age(r.get("appt_dt"), _now_ct)
                delta_txt = f"&#9888; verify delivery (appt {_age})" if _age else "&#9888; verify delivery"
                delta_color = AMBER

            tbody_rows += (
                f"<tr style='border-bottom:1px solid {LINE};'>"
                f"<td style='padding:8px 10px;font-weight:700;'>{r['truck_name']}</td>"
                f"<td style='padding:8px 10px;color:{MUTE};'>{r.get('driver_name') or '—'}</td>"
                f"<td style='padding:8px 10px;'>{r['shipper'] or '—'}</td>"
                f"<td style='padding:8px 10px;color:{MUTE};'>{shipper_loc or '—'}</td>"
                f"<td style='padding:8px 10px;'>{r['consignee'] or '—'}</td>"
                f"<td style='padding:8px 10px;color:{MUTE};'>{consignee_loc or '—'}</td>"
                f"<td style='padding:8px 10px;white-space:nowrap;'>{_fmt_appt_cell(r)}</td>"
                f"<td style='padding:8px 10px;white-space:nowrap;'>{eta_str}</td>"
                f"<td style='padding:8px 10px;color:{delta_color};font-weight:700;white-space:nowrap;'>{delta_txt}</td>"
                f"<td style='padding:8px 10px;color:{hos_color};white-space:nowrap;font-weight:{'700' if r.get('hos_delay') else '400'};'>"
                f"{hos_txt}{'*' if r.get('hos_delay') else ''}</td>"
                f"<td style='padding:8px 10px;color:{MUTE};'>{r['broker'] or '—'}</td>"
                f"</tr>"
            )
        body = (
            f"<table cellpadding='0' cellspacing='0' style='width:100%;border-collapse:collapse;'>"
            f"{thead}<tbody>{tbody_rows}</tbody></table>"
        )

    # Untracked loads section (active loads with no GPS / no truck assigned)
    if untracked:
        ut_thead = (
            f"<thead><tr style='background:{TILEBG};border-bottom:1px solid {INK};'>"
            + "".join(
                f"<th style='padding:6px 10px;text-align:left;font-size:10px;"
                f"text-transform:uppercase;letter-spacing:0.8px;color:{MUTE};'>{h}</th>"
                for h in ("Load #", "Truck", "Driver", "Consignee", "City", "Appt", "Reason"))
            + "</tr></thead>"
        )
        ut_rows = ""
        for r in untracked:
            ut_city = (f"{r.get('consignee_city', '')}, {r.get('consignee_state', '')}"
                       .strip(", ") or "—")
            ut_rows += (
                f"<tr style='border-bottom:1px solid {LINE};'>"
                f"<td style='padding:6px 10px;font-weight:700;'>{r['load_no'] or '—'}</td>"
                f"<td style='padding:6px 10px;'>{r.get('truck') or '—'}</td>"
                f"<td style='padding:6px 10px;color:{MUTE};'>{r.get('driver') or '—'}</td>"
                f"<td style='padding:6px 10px;'>{r.get('consignee') or '—'}</td>"
                f"<td style='padding:6px 10px;color:{MUTE};'>{ut_city}</td>"
                f"<td style='padding:6px 10px;white-space:nowrap;'>"
                f"{_fmt_dt_ct(r.get('appt_dt')) or '—'}</td>"
                f"<td style='padding:6px 10px;color:{AMBER};font-size:12px;'>"
                f"{r.get('reason', '—')}</td>"
                f"</tr>"
            )
        untracked_section = (
            f"<div style='padding:16px 24px 4px;border-top:2px solid {LINE};'>"
            f"<div style='font-weight:700;text-transform:uppercase;font-size:11px;"
            f"letter-spacing:0.8px;color:{MUTE};'>Untracked Loads ({len(untracked)})</div>"
            f"<div style='color:{MUTE};font-size:11px;margin-top:2px;'>"
            f"Active X-Trux loads with no ETA — GPS not reporting or truck not yet assigned"
            f"</div></div>"
            f"<div style='padding:0 24px 20px;'>"
            f"<table cellpadding='0' cellspacing='0' style='width:100%;border-collapse:collapse;'>"
            f"{ut_thead}<tbody>{ut_rows}</tbody></table></div>"
        )
    else:
        untracked_section = ""

    gen_iso = generated_at.strftime("%Y-%m-%dT%H:%M:%SZ")
    staleness_js = (
        f"<script>(function(){{"
        f"var t=new Date('{gen_iso}');"
        f"function upd(){{var m=Math.round((Date.now()-t)/60000);"
        f"document.getElementById('age').textContent=m<1?'just now':m+' min ago';}};"
        f"upd();setInterval(upd,30000);}})();</script>"
    )

    return (
        "<!doctype html><html><head><meta charset='utf-8'>"
        "<meta http-equiv='refresh' content='1800'>"
        "<meta name='viewport' content='width=device-width,initial-scale=1'>"
        f"<style>body{{margin:0;background:#fff;{FONT}}}</style>"
        "</head><body>"
        f"<div style='background:{strip_bg};color:#fff;padding:8px 24px;"
        f"font-size:12px;font-weight:700;letter-spacing:0.3px;'>"
        f"{strip_msg} &middot; {len(rows)} tracked{stale_badge}{untracked_badge}</div>"
        f"<div style='padding:20px 24px;border-bottom:3px solid {RED};'>"
        f"<div style='font-weight:700;letter-spacing:1.5px;font-size:11px;"
        f"color:{RED};text-transform:uppercase;'>XFreight &middot; ETAs</div>"
        f"<div style='font-size:22px;font-weight:700;margin-top:4px;'>"
        f"Active X-Trux Loads &mdash; Live ETA</div>"
        f"<div style='color:{MUTE};font-size:12px;margin-top:6px;'>"
        f"Generated {generated_at.astimezone(CT):%a %b %d, %I:%M %p} CT"
        f" &middot; <span id='age'></span>"
        f" &middot; refreshes every 30 min</div>"
        f"</div>"
        f"<div style='padding:20px 24px;'>{body}</div>"
        + untracked_section
        + f"<div style='padding:14px 24px;color:{MUTE};font-size:11px;border-top:1px solid {LINE};'>"
        f"Delta = ETA &minus; appointment. Red = &ge;45 min late &middot; "
        f"amber = late (under 45 min) &middot; green = within 30 min early. "
        f"ETA from Samsara GPS &rarr; Mapbox driving-traffic (full remaining route). "
        f"&#9888; on ETA = GPS fix &gt;{_GPS_STALE_MINUTES} min old; delta prefixed ~ is approximate. "
        f"HOS Left = driver&rsquo;s remaining legal drive time from Samsara; "
        f"* means a 10-hour mandatory rest was added to ETA.</div>"
        + staleness_js
        + "</body></html>"
    )


def _render_xlsx(rows: list[dict], generated_at: datetime) -> bytes:
    """Build an .xlsx in memory using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "ETAs"

    ws.append([f"XFreight ETAs — generated {generated_at.astimezone(CT):%a %b %d, %I:%M %p} CT"])
    ws.merge_cells("A1:K1")
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    headers = ["Truck", "Driver", "Shipper", "Shipper City", "Consignee", "Consignee City",
               "Appt (CT)", "ETA (CT)", "Delta (min)", "HOS Left", "Broker"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="FAFAFA")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=3, column=col_idx)
        c.font = Font(bold=True)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left")

    for r in rows:
        hos_cell = _fmt_hos(r.get("hos_remaining_s"))
        if r.get("hos_delay"):
            hos_cell += " (+ 10h rest)"
        ws.append([
            r["truck_name"],
            r.get("driver_name") or "",
            r["shipper"],
            f"{r['shipper_city']}, {r['shipper_state']}".strip(", "),
            r["consignee"],
            f"{r['consignee_city']}, {r['consignee_state']}".strip(", "),
            (f"{_fmt_dt_ct(r['appt_window_begin_dt'])} – {_fmt_dt_ct(r['appt_dt'])}"
             if r.get("appt_window_begin_dt") else _fmt_dt_ct(r["appt_dt"])),
            _fmt_dt_ct(r.get("eta_dt")),
            r.get("delta_min", ""),
            hos_cell,
            r["broker"],
        ])

    widths = [10, 22, 28, 22, 28, 22, 22, 22, 14, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )
    load_dotenv()

    mapbox_token = os.environ.get("MAPBOX_TOKEN")
    if not mapbox_token:
        log.error("MAPBOX_TOKEN not set — aborting.")
        return 1

    tenant_id = os.environ.get("AZURE_TENANT_ID")
    client_id = os.environ.get("AZURE_CLIENT_ID")
    client_secret = os.environ.get("AZURE_CLIENT_SECRET")
    user_upn = os.environ.get("ONEDRIVE_USER_UPN")
    if not all([tenant_id, client_id, client_secret, user_upn]):
        log.error("Azure / OneDrive env vars missing — aborting.")
        return 1

    # --- Pull active loads from Alvys -----------------------------------
    alvys = AlvysClient(
        client_id=os.environ["ALVYS_CLIENT_ID"],
        client_secret=os.environ["ALVYS_CLIENT_SECRET"],
    )
    trucks = alvys.fetch_trucks()
    trucks_by_id = {}
    for t in trucks:
        tid = str(t.get("Id") or "")
        if not tid:
            continue
        num = (t.get("TruckNum") or t.get("TruckNumber")
               or t.get("Number") or t.get("Name") or "")
        trucks_by_id[tid] = str(num)
    log.info("Indexed %d trucks", len(trucks_by_id))

    raw_drivers = alvys.fetch_drivers()
    drivers_by_id = {}
    for d in raw_drivers:
        did = str(d.get("Id") or "")
        if not did:
            continue
        name = (d.get("FullName") or d.get("Name") or d.get("DisplayName") or "")
        drivers_by_id[did] = str(name)
    log.info("Indexed %d drivers", len(drivers_by_id))

    raw_users = alvys.fetch_users()
    users_by_id = {}
    for u in raw_users:
        uid = str(u.get("Id") or "")
        if not uid:
            continue
        name = (u.get("FullName") or u.get("DisplayName") or u.get("Name")
                or f"{u.get('FirstName', '')} {u.get('LastName', '')}".strip())
        if name:
            users_by_id[uid] = name
    log.info("Indexed %d users", len(users_by_id))

    # Active loads: status filter + ~7 day updatedAt window
    start_date = (datetime.now(CT) - timedelta(days=7)).strftime("%Y-%m-%d")
    all_loads = alvys.fetch_loads(start_date)
    log.info("Pulled %d total loads from last 7d", len(all_loads))

    active = [
        L for L in all_loads
        if L.get("Status") in ACTIVE_STATUSES
        # Office identity comes from InvoiceAs (free string like "X-TRUX INC");
        # OfficeId would require the lookups module, which we skip for speed.
        and _entity_is_xtrux(L.get("InvoiceAs"))
    ]
    log.info("Filtered to %d active X-Trux loads", len(active))

    # Trips carry truck assignment — loads don't embed it.
    # Use fetch_active_trips (no date filter) so long-haul loads whose truck was
    # assigned >7 days ago but whose appointment changed today are still included.
    raw_trips = alvys.fetch_active_trips()
    trips_by_load: dict[str, dict] = {}
    for trip in raw_trips:
        ln = str(trip.get("LoadNumber") or "")
        if ln and ln not in trips_by_load:
            trips_by_load[ln] = trip
    log.info("Indexed %d trips (%d unique load numbers)", len(raw_trips), len(trips_by_load))

    load_rows: list[dict] = []
    for L in active:
        ln = str(L.get("LoadNumber") or "")
        row = _extract_load_row(L, trucks_by_id, trips_by_load, drivers_by_id, users_by_id)
        if row:
            load_rows.append(row)
        else:
            # Diagnostic: log why each active load was dropped
            next_s = _next_undelivered_stop(L)
            trip = trips_by_load.get(ln)
            truck_obj = (trip.get("Truck") or {}) if trip else {}
            coords = (next_s.get("Coordinates") or {}) if next_s else {}
            log.info("SKIP load %s: trip=%s truck_obj=%s next_stop=%s coords=%s",
                     ln, "found" if trip else "MISSING",
                     truck_obj.get("TruckNum") or truck_obj.get("Number") or truck_obj,
                     "found" if next_s else "MISSING",
                     coords)
    log.info("Routable loads (have truck + undelivered stop + dest coords): %d",
             len(load_rows))

    # Diagnostic: when we can't extract any rows, dump load + trip + truck shape.
    if not load_rows and active:
        import json
        sample = active[0]
        sample_ln = str(sample.get("LoadNumber") or "")
        sample_trip = trips_by_load.get(sample_ln)
        log.warning("=== DIAGNOSTIC: no routable loads ===")
        log.warning("Load top-level keys: %s", sorted(sample.keys()))
        stops = sample.get("Stops") or []
        log.warning("Stops count: %d", len(stops))
        if stops:
            log.warning("First stop keys: %s", sorted(stops[0].keys()))
            log.warning("First stop Address: %s",
                        json.dumps(stops[0].get("Address"), default=str)[:400])
            log.warning("First stop Coordinates: %s",
                        json.dumps(stops[0].get("Coordinates"), default=str)[:400])
        log.warning("trips_by_load has load %r? %s", sample_ln, sample_trip is not None)
        if sample_trip:
            log.warning("Trip top-level keys: %s", sorted(sample_trip.keys()))
            truck_obj = sample_trip.get("Truck")
            log.warning("Trip.Truck field: %s", json.dumps(truck_obj, default=str)[:800])
        else:
            log.warning("No matching trip — active load numbers vs trip load numbers (sample):")
            active_lns = [str(L.get("LoadNumber") or "") for L in active[:5]]
            trip_lns = list(trips_by_load.keys())[:5]
            log.warning("  active: %s  trips: %s", active_lns, trip_lns)
        log.warning("Trucks_by_id sample (first 3): %s", list(trucks_by_id.items())[:3])
        log.warning("=== END DIAGNOSTIC ===")

    # --- Pull current locations + HOS from Samsara ----------------------
    samsara = SamsaraClient(api_token=os.environ["SAMSARA_API_TOKEN"])
    locations = samsara.fetch_locations()
    locs_by_truck = _locations_by_truck_name(locations)
    log.info("Resolved current GPS for %d trucks", len(locs_by_truck))

    hos_clocks = samsara.fetch_hos_clocks()
    hos_index = _build_hos_index(hos_clocks)
    log.info("Indexed HOS remaining drive time for %d drivers", len(hos_index))

    # --- Compute ETA per load via Mapbox --------------------------------
    now = datetime.now(timezone.utc)
    rows_with_eta: list[dict] = []
    for row in load_rows:
        gps = _lookup_truck_gps(locs_by_truck, row["truck_name"])
        if not gps:
            log.info("SKIP load %s (truck %s): no Samsara GPS",
                     row["load_no"], row["truck_name"])
            continue

        # Check GPS fix age — stale positions produce unreliable ETAs.
        gps_age_min: int | None = None
        gps_stale = False
        if gps.get("ts"):
            gps_age_min = int((now - gps["ts"]).total_seconds() / 60)
            gps_stale = gps_age_min > _GPS_STALE_MINUTES
            if gps_stale:
                log.warning("load %s truck %s: GPS fix is %d min old — ETA unreliable",
                            row["load_no"], row["truck_name"], gps_age_min)

        # Full remaining route: truck → every remaining stop in order.
        duration_s = _mapbox_duration_seconds(
            mapbox_token, gps["lat"], gps["lng"],
            row["route_waypoints"],
        )
        if duration_s is None:
            log.info("SKIP load %s (truck %s): Mapbox returned no route",
                     row["load_no"], row["truck_name"])
            continue

        n_stops = len(row["route_waypoints"])
        if n_stops > 1:
            log.info("  load %s: %d remaining stop(s), full-route duration %.1fh",
                     row["load_no"], n_stops, duration_s / 3600)

        # HOS cap: if driver runs out of legal drive time before arriving,
        # add the federal 10-hour mandatory rest to the ETA.
        hos_remaining_s = _hos_remaining(hos_index, row.get("driver_name", ""))
        hos_delay_s = 0
        if hos_remaining_s is not None and duration_s > hos_remaining_s:
            hos_delay_s = _HOS_REST_SECONDS
            log.info("  load %s HOS cap: driver %s has %.1fh left, route %.1fh → +10h rest",
                     row["load_no"], row.get("driver_name"),
                     hos_remaining_s / 3600, duration_s / 3600)

        eta_dt = now + timedelta(seconds=duration_s + hos_delay_s)
        delta_min = None
        if row["appt_dt"]:
            delta_min = int(round((row["appt_dt"] - eta_dt).total_seconds() / 60))

        # Stale-appointment guard: an appt > _STALE_APPT_HOURS in the past means
        # the truck is almost certainly already delivered (ArrivedAt unset in
        # Alvys) or the appt was never rescheduled — not a live "Xh late" event.
        # Flag it so it shows as "verify delivery" instead of firing a hard alert.
        appt_stale = _is_appt_stale(row["appt_dt"], now)

        log.info("load %s truck %-6s  gps=%s  hos=%s  appt=%s  eta=%s  delta=%s min%s",
                 row["load_no"], row["truck_name"],
                 f"{gps_age_min}m old" if gps_age_min is not None else "fresh",
                 _fmt_hos(hos_remaining_s),
                 _fmt_dt_ct(row["appt_dt"]), _fmt_dt_ct(eta_dt), delta_min,
                 f"  [STALE APPT {_fmt_appt_age(row['appt_dt'], now)} — verify delivery]"
                 if appt_stale else "")
        rows_with_eta.append({
            **row,
            "eta_dt": eta_dt,
            "delta_min": delta_min,
            "hos_remaining_s": hos_remaining_s,
            "hos_delay": hos_delay_s > 0,
            "gps_age_min": gps_age_min,
            "gps_stale": gps_stale,
            "appt_stale": appt_stale,
        })

    log.info("Computed ETAs for %d active loads", len(rows_with_eta))

    # --- Build untracked list for the HTML "Untracked Loads" section ----
    tracked_load_nos = {str(r["load_no"]) for r in rows_with_eta}
    load_rows_by_ln = {str(r["load_no"]): r for r in load_rows}
    untracked: list[dict] = []
    for L in active:
        ln = str(L.get("LoadNumber") or "")
        if ln in tracked_load_nos:
            continue
        stops = L.get("Stops") or []
        last_stop = stops[-1] if stops else {}
        trip = trips_by_load.get(ln)
        truck_obj = (trip.get("Truck") or {}) if trip else {}
        truck = (truck_obj.get("TruckNum") or truck_obj.get("TruckNumber")
                 or truck_obj.get("Number") or truck_obj.get("Name") or "")
        driver_obj = (trip.get("Driver1") or {}) if trip else {}
        driver = (driver_obj.get("FullName") or driver_obj.get("Name") or "")
        if ln in load_rows_by_ln:
            row_r = load_rows_by_ln[ln]
            reason = ("No GPS fix" if not _lookup_truck_gps(locs_by_truck, row_r["truck_name"])
                      else "No Mapbox route")
        elif not trip:
            reason = "No trip found"
        elif not truck:
            reason = "No truck assigned"
        else:
            reason = "No stop coordinates"
        untracked.append({
            "load_no": ln,
            "truck": truck,
            "driver": driver,
            "consignee": _g(last_stop, "CompanyName") or "",
            "consignee_city": _g(last_stop, "Address", "City") or "",
            "consignee_state": _g(last_stop, "Address", "State") or "",
            "appt_dt": _parse_iso(_stop_appt_iso(last_stop)),
            "reason": reason,
        })
    log.info("Untracked active loads (no ETA computed): %d", len(untracked))

    # --- OneDrive token + folder (needed for Teams state AND file upload) ---
    folder = os.environ.get("ETA_ONEDRIVE_FOLDER", "ETA").strip("/")
    tok = get_token(tenant_id, client_id, client_secret)
    ensure_folder(tok, user_upn, folder)

    # --- Teams alert for loads 45+ min late -----------------------------
    # Exclude stale GPS (ETA may be wrong) AND stale appointments (the load is
    # almost certainly already delivered / never rescheduled — a hard "153h late"
    # card would just be alert noise on an already-serviced stop).
    late = [r for r in rows_with_eta
            if r.get("delta_min") is not None
            and r["delta_min"] <= _LATE_THRESHOLD_MIN
            and not r.get("gps_stale")
            and not r.get("appt_stale")]
    stale_appt_rows = [r for r in rows_with_eta if r.get("appt_stale")]
    if stale_appt_rows:
        log.info("Suppressed %d stale-appt load(s) from the Teams late alert "
                 "(appt > %dh old — verify delivery/reschedule): %s",
                 len(stale_appt_rows), _STALE_APPT_HOURS,
                 ", ".join(f"#{r['load_no']} ({_fmt_appt_age(r['appt_dt'], now)})"
                           for r in stale_appt_rows))
    webhook = os.environ.get("TEAMS_ETA_WEBHOOK", "").strip()
    if webhook:
        _sync_teams_webhook(webhook, tok, user_upn, folder, late)
    else:
        log.info("TEAMS_ETA_WEBHOOK not set — skipping Teams alert")

    # --- Render + upload ------------------------------------------------
    generated_at = datetime.now(timezone.utc)
    html = _render_html(rows_with_eta, generated_at, untracked)
    xlsx_bytes = _render_xlsx(rows_with_eta, generated_at)

    out_dir = Path("output/eta")
    out_dir.mkdir(parents=True, exist_ok=True)
    html_path = out_dir / "XFreight_ETAs.html"
    xlsx_path = out_dir / "XFreight_ETAs.xlsx"
    html_path.write_text(html, encoding="utf-8")
    xlsx_path.write_bytes(xlsx_bytes)

    upload_file(tok, user_upn, folder, html_path.name, html_path)
    upload_file(tok, user_upn, folder, xlsx_path.name, xlsx_path)
    log.info("Published %s/XFreight_ETAs.{html,xlsx} to OneDrive", folder)

    # --- Append to rolling 90-day ETA log on OneDrive -------------------
    _append_eta_log(tok, user_upn, folder, generated_at, rows_with_eta, untracked)

    return 0


if __name__ == "__main__":
    sys.exit(main())
