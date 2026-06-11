"""
Post-process the Alvys TMS export before uploading to OneDrive.

The Alvys TMS now puts brokered (X-Linx) load costs in "Carrier Rate"
instead of "Driver Rate".  Power BI and the scorecard email both use
"Driver Rate" as the cost column, so those loads appear at ~100% margin.

This script:
  1. Reads the exported Excel file (default: "Alvys Master 2026.xlsx")
  2. On the Loads sheet: for any row where Driver Rate = 0 and
     Carrier Rate > 0, sets Driver Rate = Carrier Rate
  3. Recomputes Gross Margin = Customer Revenue - Driver Rate
  4. Saves the fixed file (overwrites in place, or to a new path)

Run AFTER downloading the TMS export, BEFORE uploading to OneDrive:

    python -m src.fix_master_driver_rate "Alvys Master 2026.xlsx"

Or to save to a different output path:

    python -m src.fix_master_driver_rate "Alvys Master 2026.xlsx" --out "Alvys Master 2026 Fixed.xlsx"

The script is safe to re-run — rows that already have Driver Rate > 0
(X-Trux / company-driver loads) are left untouched.
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("fix_master_driver_rate")

LOADS_SHEET = "Loads"


def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Return the first column name from candidates that exists in df."""
    cols_lower = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c in df.columns:
            return c
        if c.lower() in cols_lower:
            return cols_lower[c.lower()]
    return None


def fix_driver_rate(path: str | Path, out_path: str | Path | None = None) -> int:
    """Read the Excel file, fix Driver Rate on the Loads sheet, write back.

    Returns the number of rows patched.
    """
    path = Path(path)
    out_path = Path(out_path) if out_path else path

    if not path.exists():
        log.error("File not found: %s", path)
        sys.exit(1)

    log.info("Reading %s …", path)
    # Read with openpyxl engine to preserve all sheets + formatting
    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet_names = xl.sheet_names
    log.info("Sheets found: %s", sheet_names)

    loads_name = next((s for s in sheet_names if s.strip().lower() == "loads"), None)
    if not loads_name:
        log.error("No 'Loads' sheet found in %s — sheets are: %s", path, sheet_names)
        sys.exit(1)

    loads = xl.parse(loads_name)
    log.info("Loads sheet: %d rows × %d columns", len(loads), len(loads.columns))

    driver_col = _find_col(loads, ["Driver Rate", "DriverRate"])
    carrier_col = _find_col(loads, ["Carrier Rate", "CarrierRate"])
    revenue_col = _find_col(loads, ["Customer Revenue", "CustomerRevenue", "Revenue"])
    margin_col  = _find_col(loads, ["Gross Margin", "GrossMargin", "Margin"])

    if not driver_col:
        log.error("No 'Driver Rate' column found. Columns: %s", list(loads.columns))
        sys.exit(1)
    if not carrier_col:
        log.warning("No 'Carrier Rate' column found — nothing to fix.")
        return 0

    log.info("Driver Rate column  : %r", driver_col)
    log.info("Carrier Rate column : %r", carrier_col)
    log.info("Revenue column      : %r", revenue_col)
    log.info("Gross Margin column : %r", margin_col)

    driver  = pd.to_numeric(loads[driver_col],  errors="coerce").fillna(0)
    carrier = pd.to_numeric(loads[carrier_col], errors="coerce").fillna(0)

    # Rows to patch: Driver Rate = 0 and Carrier Rate > 0
    patch_mask = (driver == 0) & (carrier > 0)
    n_patched = int(patch_mask.sum())
    log.info("Rows to patch (Driver Rate=0, Carrier Rate>0): %d of %d",
             n_patched, len(loads))

    if n_patched == 0:
        log.info("Nothing to fix — Driver Rate already populated for all rows with cost.")
        return 0

    # Apply the fix: copy Carrier Rate → Driver Rate for brokered loads
    loads[driver_col] = driver.where(~patch_mask, carrier)

    # Recompute Gross Margin if the column exists
    if margin_col and revenue_col:
        revenue = pd.to_numeric(loads[revenue_col], errors="coerce").fillna(0)
        new_driver = pd.to_numeric(loads[driver_col], errors="coerce").fillna(0)
        loads[margin_col] = revenue - new_driver
        log.info("Recomputed %r = %r - %r", margin_col, revenue_col, driver_col)
    elif margin_col:
        log.warning("Revenue column not found — Gross Margin not recomputed.")

    # Write back: use openpyxl to preserve all other sheets
    log.info("Writing fixed file → %s …", out_path)
    wb = load_workbook(path)
    ws = wb[loads_name]

    # Find the column indices in the worksheet (row 1 = headers)
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {name: i + 1 for i, name in enumerate(header_row) if name is not None}

    dr_col_idx = col_index.get(driver_col)
    gm_col_idx = col_index.get(margin_col) if margin_col else None

    if dr_col_idx is None:
        log.error("Could not find %r in worksheet header row", driver_col)
        sys.exit(1)

    patched_rows = loads.index[patch_mask].tolist()
    for df_row_idx in patched_rows:
        ws_row = df_row_idx + 2  # +1 for 1-based, +1 for header row
        ws.cell(row=ws_row, column=dr_col_idx).value = float(loads.at[df_row_idx, driver_col])
        if gm_col_idx and revenue_col:
            ws.cell(row=ws_row, column=gm_col_idx).value = float(loads.at[df_row_idx, margin_col])

    wb.save(out_path)
    log.info("Saved. %d rows patched.", n_patched)
    return n_patched


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fix Driver Rate in Alvys Master Excel before OneDrive upload."
    )
    parser.add_argument("file", help="Path to the Alvys TMS export Excel file")
    parser.add_argument("--out", default=None,
                        help="Output path (default: overwrite input file)")
    args = parser.parse_args()

    n = fix_driver_rate(args.file, args.out)
    if n > 0:
        out = args.out or args.file
        print(f"\n✓ Patched {n} X-Linx load(s). File ready to upload: {out}")
    else:
        print("\n✓ No rows needed patching.")


if __name__ == "__main__":
    main()
